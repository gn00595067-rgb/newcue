# -*- coding: utf-8 -*-
"""
簡易模式 · 子公司忠實渲染器 (Subsidiary faithful renderer)

依 cueexample 的「Media Schedule」子公司範本，把 calculate_plan_data 的計算結果
渲染成一檔多分頁（每個秒數一版，客戶挑秒數）。相對舊東吳格式，加入：
  - 區域全名（北區→北區-北北基…，用 config.REGION_DISPLAY_MAP）
  - 家樂福改名：萬家福（量販）/ 樂家康（超市）
  - 加強欄：總檔次、總曝光次數(店數×總檔次)、曝光期間店舖總人流量(曝光×係數)
  - 底部 Total / 製作 / 5% VAT / Grand Total，週末日期淡黃標記
本模組不 import streamlit，可獨立測試。
"""
import io
from datetime import timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import config

# 曝光期間店舖總人流量 = 總曝光次數 × 此係數（由範本反推固定值 125/18≈6.9444）
TRAFFIC_FACTOR = 125.0 / 18.0

# 子公司平台顯示（Station 欄）
STATION_NAME = {
    "全家廣播": "全家便利商店\n通路廣播廣告",
    "新鮮視":   "全家便利商店\n全家新鮮視(全家電視)",
}

FONT = "微軟正黑體"
_thin = Side(style="thin", color="808080")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
WEEKEND_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
MONEY = '"$"#,##0'
WEEKDAY_CH = ["一", "二", "三", "四", "五", "六", "日"]


def _c(ws, r, col, val, *, bold=False, align=CENTER, fill=None, border=True, num=None, size=10):
    cell = ws.cell(row=r, column=col, value=val)
    cell.font = Font(name=FONT, bold=bold, size=size)
    cell.alignment = align
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    if num:
        cell.number_format = num
    return cell


def _region_disp(region):
    return config.REGION_DISPLAY_MAP.get(region, region)


def _block_rows(rows, media):
    return [r for r in rows if r.get("media") == media]


def _render_sheet(wb, title, second, rows, budget, prod_cost, start_dt, end_dt,
                  client, tax_id, product, medium_label, media_order, remarks):
    ws = wb.create_sheet(title=title[:31])
    days = (end_dt - start_dt).days + 1
    C_STA, C_LOC, C_PRG, C_DAY, C_SIZE, C_RATE, C_PKG = 1, 2, 3, 4, 5, 6, 7
    C_DAY0 = 8
    C_TOT = C_DAY0 + days
    C_EXP = C_TOT + 1
    C_TRA = C_TOT + 2
    last_col = C_TRA

    # --- 抬頭 ---
    _c(ws, 1, 1, "Media Schedule", bold=True, align=Alignment(horizontal="center", vertical="center"),
       border=False, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=C_PKG)
    _c(ws, 3, 1, "客戶名稱：", align=LEFT, border=False)
    _c(ws, 3, 2, client or "", align=LEFT, border=False)
    _c(ws, 4, 1, "Product：", align=LEFT, border=False)
    _c(ws, 4, 2, (f"{product}  " if product else "") + f"{second}秒", align=LEFT, border=False)
    _c(ws, 5, 1, "Period：", align=LEFT, border=False)
    _c(ws, 5, 2, f"{start_dt.strftime('%Y. %m. %d')} - {end_dt.strftime('%Y. %m. %d')}",
       align=LEFT, border=False)
    _c(ws, 6, 1, "Medium：", align=LEFT, border=False)
    _c(ws, 6, 2, medium_label, align=LEFT, border=False)

    # --- 欄位標題 (row7 主標，row8 星期) ---
    HR = 7
    heads = {C_STA: "Station", C_LOC: "Location", C_PRG: "Program", C_DAY: "Day-part",
             C_SIZE: "Size", C_RATE: "rate (Net)", C_PKG: "Package-cost\n(Net)"}
    for col, txt in heads.items():
        _c(ws, HR, col, txt, bold=True, fill=HEAD_FILL)
        ws.merge_cells(start_row=HR, start_column=col, end_row=HR + 1, end_column=col)
    # 每日日期 + 星期
    d = start_dt
    for i in range(days):
        col = C_DAY0 + i
        weekend = d.weekday() >= 5
        _c(ws, HR, col, d.strftime("%m/%d"), bold=True, fill=WEEKEND_FILL if weekend else HEAD_FILL)
        _c(ws, HR + 1, col, WEEKDAY_CH[d.weekday()], fill=WEEKEND_FILL if weekend else HEAD_FILL)
        d += timedelta(days=1)
    for col, txt in {C_TOT: "總檔次", C_EXP: "總曝光次數", C_TRA: "曝光期間\n店舖總人流量"}.items():
        _c(ws, HR, col, txt, bold=True, fill=HEAD_FILL)
        ws.merge_cells(start_row=HR, start_column=col, end_row=HR + 1, end_column=col)

    # --- 資料列 ---
    r = HR + 2
    first_data = True
    exp_total = 0
    tra_total = 0.0
    for media in media_order:
        brows = _block_rows(rows, media)
        if not brows:
            continue
        block_start = r
        if media == "家樂福":
            # 萬家福量販 + 樂家康超市
            for row in brows:
                is_super = "超市" in str(row.get("region", ""))
                store = int(row.get("program_num", 0))
                spots = int(row.get("spots", 0))
                sch = row.get("schedule", [])
                _c(ws, r, C_STA, "全省" if r == block_start else "", align=CENTER)
                _c(ws, r, C_LOC, "樂家康-超市" if is_super else "萬家福-量販店", align=CENTER)
                _c(ws, r, C_PRG, f"{store}店")
                _c(ws, r, C_DAY, row.get("daypart", ""))
                _c(ws, r, C_SIZE, f"{second}秒")
                rate = row.get("rate_display", "")
                _c(ws, r, C_RATE, rate, num=MONEY if isinstance(rate, (int, float)) else None)
                if first_data:
                    _c(ws, r, C_PKG, round(float(budget)), num=MONEY)
                    first_data = False
                else:
                    _c(ws, r, C_PKG, "")
                for i in range(days):
                    _c(ws, r, C_DAY0 + i, sch[i] if i < len(sch) else 0)
                exp = store * spots
                tra = exp * TRAFFIC_FACTOR
                _c(ws, r, C_TOT, spots)
                _c(ws, r, C_EXP, exp, num="#,##0")
                _c(ws, r, C_TRA, round(tra), num="#,##0")
                exp_total += exp
                tra_total += tra
                r += 1
        else:
            for row in brows:
                store = int(row.get("program_num", 0))
                spots = int(row.get("spots", 0))
                sch = row.get("schedule", [])
                _c(ws, r, C_STA, STATION_NAME.get(media, media) if r == block_start else "", align=CENTER)
                _c(ws, r, C_LOC, _region_disp(row.get("region", "")))
                _c(ws, r, C_PRG, store)
                _c(ws, r, C_DAY, row.get("daypart", "") if r == block_start else "")
                _c(ws, r, C_SIZE, f"{second}秒" if r == block_start else "")
                _c(ws, r, C_RATE, row.get("rate_display", ""), num=MONEY)
                if first_data:
                    _c(ws, r, C_PKG, round(float(budget)), num=MONEY)
                    first_data = False
                else:
                    _c(ws, r, C_PKG, "")
                for i in range(days):
                    _c(ws, r, C_DAY0 + i, sch[i] if i < len(sch) else 0)
                exp = store * spots
                tra = exp * TRAFFIC_FACTOR
                _c(ws, r, C_TOT, spots)
                _c(ws, r, C_EXP, exp, num="#,##0")
                _c(ws, r, C_TRA, round(tra), num="#,##0")
                exp_total += exp
                tra_total += tra
                r += 1
        # 合併 Station 欄
        if r - 1 > block_start:
            ws.merge_cells(start_row=block_start, start_column=C_STA, end_row=r - 1, end_column=C_STA)

    # --- 費用區 ---
    vat = round(float(budget) * 0.05)
    grand = round(float(budget) + float(prod_cost) + vat)
    def fee(label, val, r_):
        _c(ws, r_, C_SIZE, label, bold=True, fill=TOTAL_FILL)
        _c(ws, r_, C_RATE, "", fill=TOTAL_FILL)
        _c(ws, r_, C_PKG, val, bold=True, num=MONEY, fill=TOTAL_FILL)
    _c(ws, r, C_SIZE, "Total", bold=True, fill=TOTAL_FILL)
    _c(ws, r, C_RATE, "", fill=TOTAL_FILL)
    _c(ws, r, C_PKG, round(float(budget)), bold=True, num=MONEY, fill=TOTAL_FILL)
    _c(ws, r, C_EXP, exp_total, bold=True, num="#,##0", fill=TOTAL_FILL)
    _c(ws, r, C_TRA, round(tra_total), bold=True, num="#,##0", fill=TOTAL_FILL)
    r += 1
    fee("製作", round(float(prod_cost)), r); r += 1
    fee("5% VAT", vat, r); r += 1
    fee("Grand Total", grand, r); r += 1

    # --- 備註 ---
    r += 1
    _c(ws, r, 1, "Remarks：", bold=True, align=LEFT, border=False); r += 1
    for line in remarks:
        _c(ws, r, 1, line, align=LEFT, border=False)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        r += 1

    # --- 欄寬 ---
    widths = {C_STA: 20, C_LOC: 14, C_PRG: 8, C_DAY: 11, C_SIZE: 7, C_RATE: 11, C_PKG: 13,
              C_TOT: 8, C_EXP: 12, C_TRA: 15}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for i in range(days):
        ws.column_dimensions[get_column_letter(C_DAY0 + i)].width = 5
    ws.row_dimensions[1].height = 24
    return ws


def render_subsidiary_workbook(combo, seconds_rows, budget, prod_cost, start_dt, end_dt,
                               client, tax_id, product, remarks):
    """
    seconds_rows: dict {second: rows}（各秒數已算好的 calculate_plan_data 結果）
    combo: 需含 "media"(list) 與 "medium"(str, Medium 欄顯示)
    回傳 xlsx bytes（一檔多分頁，每秒數一版）。
    """
    wb = Workbook()
    wb.remove(wb.active)
    for second in combo["seconds"]:
        rows = seconds_rows[second]
        _render_sheet(wb, f"{second}秒版", second, rows, budget, prod_cost,
                      start_dt, end_dt, client, tax_id, product,
                      combo.get("medium", ""), combo["media"], remarks)
    if not wb.worksheets:
        wb.create_sheet("空")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
