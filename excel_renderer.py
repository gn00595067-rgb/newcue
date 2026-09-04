"""
Excel 報表生成引擎 (Excel Renderer)
負責生成三種格式的 Excel 報表：東吳、聲活、鉑霖
"""

import streamlit as st
import io
from datetime import timedelta
import math
import unicodedata
from decimal import Decimal, ROUND_HALF_UP
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from config import (
    FONT_MAIN, BS_THIN, BS_MEDIUM, BS_HAIR, FMT_MONEY, FMT_NUMBER
)
from pdf_converter import get_cloud_logo_bytes
from utils import split_period_by_months


def _round_half_up(value):
    return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


@st.cache_data(show_spinner="正在生成 Excel 報表...", ttl=3600)
def generate_excel_from_scratch(format_type, start_dt, end_dt, client_name, tax_id, product_name, rows, remarks_list, final_budget_val, prod_cost, sales_person, total_list_accum=None):
    """
    主函式：根據格式類型生成 Excel 報表

    參數:
        format_type: "東吳" | "聲活" | "鉑霖"
        start_dt, end_dt: 日期範圍
        client_name, tax_id, product_name: 客戶資訊
        rows: 計算後的排程資料
        remarks_list: 備註文字列表
        final_budget_val: 最終預算（Package-cost Total 用此顯示，讓客戶有折扣感）
        total_list_accum: 未使用，保留參數相容；檔次計算以實作價為準
        prod_cost: 製作費
        sales_person: 業務名稱
    """
    if total_list_accum is None:
        total_list_accum = final_budget_val
    
    # 定義 Excel 通用樣式 (邊框、字體、對齊)
    SIDE_THIN, SIDE_MEDIUM, SIDE_HAIR = Side(style=BS_THIN), Side(style=BS_MEDIUM), Side(style=BS_HAIR)
    SIDE_DOUBLE = Side(style='double')
    BORDER_ALL_THIN = Border(top=SIDE_THIN, bottom=SIDE_THIN, left=SIDE_THIN, right=SIDE_THIN)
    BORDER_ALL_MEDIUM = Border(top=SIDE_MEDIUM, bottom=SIDE_MEDIUM, left=SIDE_MEDIUM, right=SIDE_MEDIUM)
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center', wrap_text=True)
    FONT_STD, FONT_BOLD, FONT_TITLE = Font(name=FONT_MAIN, size=12), Font(name=FONT_MAIN, size=14, bold=True), Font(name=FONT_MAIN, size=48, bold=True)
    FONT_HEADER, FONT_DAILY, FONT_REMARKS, FONT_SIGN = Font(name=FONT_MAIN, size=20), Font(name=FONT_MAIN, size=16), Font(name=FONT_MAIN, size=18), Font(name=FONT_MAIN, size=20)
    FONT_16, FONT_16_BOLD = Font(name=FONT_MAIN, size=16), Font(name=FONT_MAIN, size=16, bold=True)
    FONT_WEEKEND = Font(name=FONT_MAIN, size=16)
    FILL_WEEKEND = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    
    # 邊框設定 Helper
    def set_border(cell, top=None, bottom=None, left=None, right=None):
        cur = cell.border
        new_top = Side(style=top) if top else cur.top
        new_bottom = Side(style=bottom) if bottom else cur.bottom
        new_left = Side(style=left) if left else cur.left
        new_right = Side(style=right) if right else cur.right
        cell.border = Border(top=new_top, bottom=new_bottom, left=new_left, right=new_right)

    # 快速繪製外框 Helper
    def draw_outer_border_fast(ws, min_r, max_r, min_c, max_c):
        for c in range(min_c, max_c + 1):
            set_border(ws.cell(min_r, c), top=BS_MEDIUM)
            set_border(ws.cell(max_r, c), bottom=BS_MEDIUM)
        for r in range(min_r, max_r + 1):
            set_border(ws.cell(r, min_c), left=BS_MEDIUM)
            set_border(ws.cell(r, max_c), right=BS_MEDIUM)

    def calc_remark_row_height(text, font_size=18, min_height=25, chars_per_line=52):
        """依備註字數粗估列高，避免長句被截字（含手動換行）。"""
        t = str(text or "")
        segs = t.split("\n") if "\n" in t else [t]
        visual_lines = 0
        for seg in segs:
            visual_lines += max(1, math.ceil(len(seg) / max(1, chars_per_line)))
        # CJK 在 LibreOffice 匯出 PDF 時列高常偏小，額外保留 padding 避免截字
        est = int(visual_lines * (font_size + 12) + 4)
        return max(min_height, est)

    # ---------------------------------------------------------
    # Sub-Engine: Dongwu (東吳格式)
    # start_row: 此區塊起始列 (1-based)。skip_footer: 若 True 不畫 Remarks/簽名。
    # first_block: 若 True 畫大標題（僅第一區塊）。skip_title: 若 True 不畫大標題（第二區塊起）。
    # ---------------------------------------------------------
    def render_dongwu_optimized(ws, start_dt, end_dt, rows, budget, prod, pkg_total=None, start_row=1, skip_footer=False, first_block=False, skip_title=False):
        if skip_title:
            base = start_row - 3
            def R(r): return base + r
        else:
            base = start_row - 1
            def R(r): return base + r
        eff_days = (end_dt - start_dt).days + 1
        spots_col_idx = 7 + eff_days + 1
        total_cols = spots_col_idx

        # 設定欄寬列高
        COL_WIDTHS = {'A': 19.6, 'B': 22.8, 'C': 14.6, 'D': 20.0, 'E': 13.0, 'F': 19.6, 'G': 17.9}
        ROW_HEIGHTS = {1: 61.0, 2: 29.0, 3: 40.0, 4: 40.0, 5: 40.0, 6: 40.0, 7: 40.0, 8: 40.0}
        
        for k, v in COL_WIDTHS.items(): ws.column_dimensions[k].width = v
        for i in range(eff_days): ws.column_dimensions[get_column_letter(8+i)].width = 8.5
        ws.column_dimensions[get_column_letter(spots_col_idx)].width = 13.0
        for r, h in ROW_HEIGHTS.items():
            if skip_title and r <= 2: continue
            ws.row_dimensions[R(r)].height = h

        # 大標題（僅當未 skip_title 時）
        if not skip_title:
            ws.merge_cells(f"A{R(1)}:{get_column_letter(total_cols)}{R(1)}"); c = ws.cell(R(1), 1); c.value = "Media Schedule"; c.font = FONT_TITLE; c.alignment = ALIGN_CENTER
        unique_secs = sorted(list(set([r['seconds'] for r in rows])))
        p_str = f"{'、'.join([f'{s}秒' for s in unique_secs])} {product_name}"
        unique_media = sorted(list(set([r['media'] for r in rows])))
        medium_str = "/".join(unique_media)
        
        infos = [("A3", "客戶名稱：", client_name), ("A4", "Product：", p_str), ("A5", "Period :", f"{start_dt.strftime('%Y. %m. %d')} - {end_dt.strftime('%Y. %m. %d')}"), ("A6", "Medium :", medium_str)]
        pos_to_row = {"A3": 3, "A4": 4, "A5": 5, "A6": 6}
        FONT_20 = Font(name=FONT_MAIN, size=20)
        for pos, lbl, val in infos:
            r_num = R(pos_to_row.get(pos, 3))
            c = ws.cell(r_num, 1); c.value = lbl; c.font = FONT_20; c.alignment = Alignment(vertical='center')
            c2 = ws.cell(c.row, 2); c2.value = val; c2.font = FONT_20; c2.alignment = Alignment(vertical='center')
        
        for c_idx in range(1, total_cols + 1): set_border(ws.cell(R(3), c_idx), top=BS_MEDIUM)
        # 日期列上方依「日曆月」分組顯示「X月」（跨月時 4月、5月 等都會顯示）
        month_groups_dw = []
        for i in range(eff_days):
            d = start_dt + timedelta(days=i)
            m_key = (d.year, d.month)
            if not month_groups_dw or month_groups_dw[-1][0] != m_key:
                month_groups_dw.append([m_key, i, i])
            else:
                month_groups_dw[-1][2] = i
        for m_key, s_idx, e_idx in month_groups_dw:
            start_col = 8 + s_idx
            end_col = 8 + e_idx
            ws.merge_cells(start_row=R(6), start_column=start_col, end_row=R(6), end_column=end_col)
            c = ws.cell(R(6), start_col)
            c.value = f"{m_key[1]}月"
            c.font = Font(name=FONT_MAIN, size=16, bold=True)
            c.alignment = ALIGN_LEFT
        
        headers = [("A","Station"), ("B","Location"), ("C","Program"), ("D","Day-part"), ("E","Size"), ("F","rate\n(Net)"), ("G","Package-cost\n(Net)")]
        for col, txt in headers:
            col_idx = column_index_from_string(col)
            ws.merge_cells(f"{col}{R(7)}:{col}{R(8)}"); c7 = ws.cell(R(7), col_idx); c7.value = txt; c8 = ws.cell(R(8), col_idx)
            c7.font = FONT_16_BOLD; c7.alignment = ALIGN_CENTER; c7.border = BORDER_ALL_THIN; c8.border = BORDER_ALL_THIN; c8.font = FONT_16
            set_border(c7, top=BS_MEDIUM); set_border(c8, bottom=BS_MEDIUM)

        # 日期標頭 (自動計算週六日底色)
        curr = start_dt
        for i in range(eff_days):
            col_idx = 8 + i
            c_d = ws.cell(R(7), col_idx); c_w = ws.cell(R(8), col_idx)
            c_d.value = curr; c_d.number_format = 'm/d'; c_w.value = ["一","二","三","四","五","六","日"][curr.weekday()]
            if curr.weekday() >= 5: c_w.fill = FILL_WEEKEND
            curr += timedelta(days=1)
            c_d.font = FONT_16; c_w.font = FONT_16; c_d.alignment = ALIGN_CENTER; c_w.alignment = ALIGN_CENTER
            c_d.border = BORDER_ALL_THIN; c_w.border = BORDER_ALL_THIN
            set_border(c_d, top=BS_MEDIUM); set_border(c_w, bottom=BS_MEDIUM)

        ws.cell(R(8), spots_col_idx - 1).border = Border(top=SIDE_THIN, bottom=SIDE_MEDIUM, left=SIDE_THIN, right=SIDE_MEDIUM)
        c_spots_7 = ws.cell(R(7), spots_col_idx); c_spots_7.value = "檔次"; c_spots_8 = ws.cell(R(8), spots_col_idx)
        ws.merge_cells(start_row=R(7), start_column=spots_col_idx, end_row=R(8), end_column=spots_col_idx)
        c_spots_7.font = FONT_16_BOLD; c_spots_7.alignment = ALIGN_CENTER; c_spots_7.border = BORDER_ALL_THIN; c_spots_8.border = BORDER_ALL_THIN; c_spots_8.font = FONT_16
        set_border(c_spots_7, top=BS_MEDIUM, left=BS_MEDIUM); set_border(c_spots_8, bottom=BS_MEDIUM, left=BS_MEDIUM)
        set_border(ws.cell(R(7), 1), right=BS_MEDIUM); set_border(ws.cell(R(8), 1), right=BS_MEDIUM)

        # 資料填充迴圈
        curr_row = R(9); grouped_data = {"全家廣播": sorted([r for r in rows if r["media"] == "全家廣播"], key=lambda x: x["seconds"]), "新鮮視": sorted([r for r in rows if r["media"] == "新鮮視"], key=lambda x: x["seconds"]), "家樂福": sorted([r for r in rows if r["media"] == "家樂福"], key=lambda x: x["seconds"])}
        total_rate_sum = 0 

        for m_key, data in grouped_data.items():
            if not data: continue
            start_merge = curr_row
            display_name = f"全家便利商店\n{m_key if m_key!='家樂福' else ''}廣告"
            if m_key == "家樂福": display_name = "家樂福"
            elif m_key == "全家廣播": display_name = "全家便利商店\n通路廣播廣告"
            elif m_key == "新鮮視": display_name = "全家便利商店\n新鮮視廣告"

            for idx, r in enumerate(data):
                ws.row_dimensions[curr_row].height = 40
                ws.cell(curr_row, 1, display_name).alignment = ALIGN_CENTER
                ws.cell(curr_row, 2, r["region"]).alignment = ALIGN_CENTER
                ws.cell(curr_row, 3, r.get("program_num", 0)).alignment = ALIGN_CENTER
                ws.cell(curr_row, 4, r["daypart"]).alignment = ALIGN_CENTER
                ws.cell(curr_row, 5, f"{r['seconds']}秒").alignment = ALIGN_CENTER
                rate = r['rate_display']; pkg = r['pkg_display']
                if isinstance(rate, (int, float)): total_rate_sum += rate
                if r.get("is_pkg_member"): pkg = r['nat_pkg_display'] if idx == 0 else None
                elif r.get("is_rebate"): pkg = r.get("pkg_display", "回饋") if (idx == 0 or not data[idx-1].get("is_rebate") or data[idx-1].get("is_bonus_rebate") != r.get("is_bonus_rebate")) else None
                elif r.get("is_custom_bonus"): pkg = r.get("pkg_display", "加贈") if (idx == 0 or not data[idx-1].get("is_custom_bonus")) else None
                c_rate = ws.cell(curr_row, 6); c_rate.value = rate; c_rate.number_format = FMT_MONEY; c_rate.alignment = ALIGN_CENTER
                if pkg is not None: c_pkg = ws.cell(curr_row, 7); c_pkg.value = pkg; c_pkg.alignment = ALIGN_CENTER; c_pkg.number_format = FMT_MONEY if isinstance(pkg, (int, float)) else '@'
                row_sum = 0
                for d_idx in range(eff_days):
                    if d_idx < len(r["schedule"]): val = r["schedule"][d_idx]; row_sum += (val if isinstance(val, (int, float)) else 0); c_s = ws.cell(curr_row, 8+d_idx); c_s.value = "" if (val == 0 or val is None) else val; c_s.number_format = FMT_NUMBER; c_s.alignment = ALIGN_CENTER; c_s.font = FONT_WEEKEND if (start_dt + timedelta(days=d_idx)).weekday() >= 5 else FONT_DAILY
                ws.cell(curr_row, spots_col_idx, row_sum).alignment = ALIGN_CENTER
                for c_idx in range(1, total_cols + 1):
                    cell = ws.cell(curr_row, c_idx); cell.border = BORDER_ALL_THIN
                    if c_idx <= 7 or c_idx == spots_col_idx: cell.font = FONT_16
                curr_row += 1

            # 合併相同媒體名稱的欄位；Column 7 依 runs (全省塊、回饋塊) 分別合併
            ws.merge_cells(start_row=start_merge, start_column=1, end_row=curr_row-1, end_column=1)
            i = 0
            while i < len(data):
                if data[i].get("is_pkg_member"):
                    j = i; 
                    while j < len(data) and data[j].get("is_pkg_member"): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=7, end_row=start_merge+j-1, end_column=7)
                    i = j
                elif data[i].get("is_rebate"):
                    j = i
                    while j < len(data) and data[j].get("is_rebate") and data[j].get("is_bonus_rebate") == data[i].get("is_bonus_rebate"): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=7, end_row=start_merge+j-1, end_column=7)
                    i = j
                elif data[i].get("is_custom_bonus"):
                    j = i
                    while j < len(data) and data[j].get("is_custom_bonus"): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=7, end_row=start_merge+j-1, end_column=7)
                    i = j
                else:
                    i += 1
            for col in [4, 5]:
                m_start = start_merge
                while m_start < curr_row:
                    m_end = m_start; val = ws.cell(m_start, col).value
                    while m_end + 1 < curr_row and ws.cell(m_end+1, col).value == val: m_end += 1
                    if m_end > m_start: ws.merge_cells(start_row=m_start, start_column=col, end_row=m_end, end_column=col)
                    m_start = m_end + 1
            draw_outer_border_fast(ws, start_merge, curr_row-1, 1, total_cols)
            for r in range(start_merge, curr_row): set_border(ws.cell(r, 1), right=BS_MEDIUM); set_border(ws.cell(r, spots_col_idx), left=BS_MEDIUM)

        # 總計列
        ws.row_dimensions[curr_row].height = 30
        c_lbl = ws.cell(curr_row, 5, "Total"); c_lbl.alignment = ALIGN_CENTER; c_lbl.font = FONT_16_BOLD
        c_rate_sum = ws.cell(curr_row, 6, total_rate_sum); c_rate_sum.number_format = FMT_MONEY; c_rate_sum.alignment = ALIGN_CENTER; c_rate_sum.font = FONT_16_BOLD
        pkg_total_val = (pkg_total if pkg_total is not None else budget)
        c_val = ws.cell(curr_row, 7, pkg_total_val); c_val.number_format = FMT_MONEY; c_val.alignment = ALIGN_CENTER; c_val.font = FONT_16_BOLD
        total_spots_all = 0
        for d_idx in range(eff_days):
            daily_sum = sum([r['schedule'][d_idx] if d_idx < len(r['schedule']) and isinstance(r['schedule'][d_idx], (int, float)) else 0 for r in rows])
            total_spots_all += daily_sum
            c = ws.cell(curr_row, 8+d_idx); c.value = "" if daily_sum == 0 else daily_sum; c.alignment = ALIGN_CENTER; c.font = FONT_WEEKEND if (start_dt + timedelta(days=d_idx)).weekday() >= 5 else FONT_DAILY; c.number_format = FMT_NUMBER
        ws.cell(curr_row, spots_col_idx, total_spots_all).alignment = ALIGN_CENTER; ws.cell(curr_row, spots_col_idx).font = FONT_16
        for c_idx in range(1, total_cols + 1): set_border(ws.cell(curr_row, c_idx), top=BS_MEDIUM, bottom=BS_MEDIUM, left=BS_THIN, right=BS_THIN)
        set_border(ws.cell(curr_row, 1), left=BS_MEDIUM, right=BS_MEDIUM); set_border(ws.cell(curr_row, spots_col_idx), left=BS_MEDIUM, right=BS_MEDIUM)
        ws.cell(curr_row, spots_col_idx - 1).border = Border(top=SIDE_MEDIUM, bottom=SIDE_MEDIUM, left=SIDE_THIN, right=SIDE_MEDIUM)
        curr_row += 1

        # 頁尾 (費用與簽名)
        vat = _round_half_up(budget * 0.05); grand_total = budget + vat
        footer_items = [("媒體", budget), ("製作", prod), ("5% VAT", vat), ("Grand Total", grand_total)]
        for label, val in footer_items:
            if label == "媒體": continue 
            ws.row_dimensions[curr_row].height = 30
            c_l = ws.cell(curr_row, 6); c_l.value = label; c_l.alignment = ALIGN_LEFT; c_l.font = FONT_16
            c_v = ws.cell(curr_row, 7); c_v.value = val; c_v.number_format = FMT_MONEY; c_v.alignment = ALIGN_CENTER; c_v.font = FONT_16
            set_border(c_l, left=BS_MEDIUM, top=BS_THIN, bottom=BS_THIN, right=BS_THIN)
            set_border(c_v, right=BS_MEDIUM, top=BS_THIN, bottom=BS_THIN, left=BS_THIN)
            if label == "Grand Total":
                set_border(ws.cell(curr_row, 6), top=BS_MEDIUM, bottom=BS_MEDIUM); set_border(ws.cell(curr_row, 7), top=BS_MEDIUM, bottom=BS_MEDIUM)
            curr_row += 1
        
        draw_outer_border_fast(ws, R(7), curr_row-1, 1, total_cols); curr_row += 1
        if skip_footer:
            return curr_row
        ws.cell(curr_row, 1, "Remarks:本排程表經雙方確認後視同合約之延伸，具同等法律約束力與效力").font = Font(name=FONT_MAIN, size=18, bold=True, underline='single')
        for rm in remarks_list:
            curr_row += 1
            is_red = rm.strip().startswith("1.") or rm.strip().startswith("4.")
            c = ws.cell(curr_row, 1); c.value = rm; c.font = Font(name=FONT_MAIN, size=18, color="FF0000" if is_red else "000000")

        curr_row += 2; sig_start = curr_row
        for _r in (sig_start, sig_start+1, sig_start+2): ws.row_dimensions[_r].height = 28
        _sig_font = Font(name=FONT_MAIN, size=20)
        ws.merge_cells(start_row=sig_start, start_column=1, end_row=sig_start, end_column=7); ws.cell(sig_start, 1, "甲    方：東吳廣告股份有限公司").alignment = ALIGN_LEFT; ws.cell(sig_start, 1).font = _sig_font
        ws.merge_cells(start_row=sig_start+1, start_column=1, end_row=sig_start+1, end_column=7); ws.cell(sig_start+1, 1, "統一編號：20935458").alignment = ALIGN_LEFT; ws.cell(sig_start+1, 1).font = _sig_font
        ws.merge_cells(start_row=sig_start+2, start_column=1, end_row=sig_start+2, end_column=7); ws.cell(sig_start+2, 1, sales_person).alignment = ALIGN_LEFT; ws.cell(sig_start+2, 1).font = _sig_font
        
        # 乙方簽名區：短天期時 total_cols 較窄，固定從 Column T 開始會太靠外（且分隔線看起來變短）。
        # 規則：<14 天時，乙方自動對齊「倒數第六欄」（即從 total_cols-5 開始，寬度 6 欄）。
        # 其他天期維持原本 8 欄寬，但若 total_cols 不足也會自動往左收。
        eff_days_for_sig = eff_days
        right_block_width = 6 if eff_days_for_sig < 14 else 8
        right_start_col = max(1, total_cols - right_block_width + 1)
        right_end_col = right_start_col + right_block_width - 1
        ws.merge_cells(start_row=sig_start, start_column=right_start_col, end_row=sig_start, end_column=right_end_col); ws.cell(sig_start, right_start_col, f"乙    方：{client_name}").alignment = ALIGN_LEFT; ws.cell(sig_start, right_start_col).font = _sig_font
        
        # 填入 Excel 統編 (東吳格式)
        ws.merge_cells(start_row=sig_start+1, start_column=right_start_col, end_row=sig_start+1, end_column=right_end_col)
        ws.cell(sig_start+1, right_start_col, f"統一編號：{tax_id}").alignment = ALIGN_LEFT; ws.cell(sig_start+1, right_start_col).font = _sig_font
        
        ws.merge_cells(start_row=sig_start+2, start_column=right_start_col, end_row=sig_start+2, end_column=right_end_col); ws.cell(sig_start+2, right_start_col, "客戶簽章：").alignment = ALIGN_LEFT; ws.cell(sig_start+2, right_start_col).font = _sig_font
        # 乙方區塊上方分隔線：固定畫到 total_cols，與上方標題底線同長
        for c_idx in range(1, total_cols + 1): set_border(ws.cell(sig_start, c_idx), top=BS_MEDIUM)
        return curr_row + 3

    # ---------------------------------------------------------
    # Sub-Engine: Shenghuo (聲活數位格式)
    # ---------------------------------------------------------
    def render_shenghuo_optimized(ws, start_dt, end_dt, rows, budget, prod, pkg_total=None):
        SIDE_DOUBLE = Side(style='double')
        eff_days = (end_dt - start_dt).days + 1
        end_c_start = 6 + eff_days
        total_cols = end_c_start + 2

        ws.column_dimensions['A'].width = 22.5; ws.column_dimensions['B'].width = 24.5; ws.column_dimensions['C'].width = 13.8; ws.column_dimensions['D'].width = 19.4; ws.column_dimensions['E'].width = 27.0
        for i in range(eff_days): ws.column_dimensions[get_column_letter(6 + i)].width = 8.1 
        ws.column_dimensions[get_column_letter(end_c_start)].width = 9.5; ws.column_dimensions[get_column_letter(end_c_start+1)].width = 58.0; ws.column_dimensions[get_column_letter(end_c_start+2)].width = 20.0 
        ROW_H_MAP = {1:30, 2:30, 3:46, 4:46, 5:40, 6:40, 7:35, 8:35}; 
        for r, h in ROW_H_MAP.items(): ws.row_dimensions[r].height = h
        
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1"); c1 = ws['A1']; c1.value = "聲活數位-媒體計劃排程表"; c1.font = Font(name=FONT_MAIN, size=24, bold=True); c1.alignment = ALIGN_CENTER
        ws.merge_cells(f"A2:{get_column_letter(total_cols)}2"); c2 = ws['A2']; c2.value = "Media Schedule"; c2.font = Font(name=FONT_MAIN, size=18, bold=True); c2.alignment = ALIGN_CENTER
        FONT_16 = Font(name=FONT_MAIN, size=16); ws.merge_cells(f"A3:{get_column_letter(total_cols)}3"); ws['A3'].value = "聲活數位科技股份有限公司 統編 28710100"; ws['A3'].font = FONT_16; ws['A3'].alignment = ALIGN_LEFT
        ws.merge_cells(f"A4:{get_column_letter(total_cols)}4"); ws['A4'].value = sales_person; ws['A4'].font = FONT_16; ws['A4'].alignment = ALIGN_LEFT
        
        unique_secs = sorted(list(set([r['seconds'] for r in rows]))); sec_str = " ".join([f"{s}秒廣告" for s in unique_secs]); period_str = f"執行期間：{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')}"
        FONT_14 = Font(name=FONT_MAIN, size=20); c5a = ws['A5']; c5a.value = "客戶名稱："; c5a.font = FONT_14; c5a.alignment = ALIGN_LEFT
        ws.merge_cells("B5:E5"); c5b = ws['B5']; c5b.value = client_name; c5b.font = FONT_14; c5b.alignment = ALIGN_LEFT
        ws.merge_cells(f"F5:{get_column_letter(end_c_start)}5"); c5f = ws['F5']; c5f.value = f"廣告規格：{sec_str}"; c5f.font = FONT_14; c5f.alignment = ALIGN_LEFT
        ws.merge_cells(f"{get_column_letter(end_c_start+1)}5:{get_column_letter(total_cols)}5"); c5_r = ws[f"{get_column_letter(end_c_start+1)}5"]; c5_r.value = period_str; c5_r.font = FONT_14; c5_r.alignment = ALIGN_LEFT 
        draw_outer_border_fast(ws, 5, 5, 1, total_cols)

        c6a = ws['A6']; c6a.value = "廣告名稱："; c6a.font = FONT_14; c6a.alignment = ALIGN_LEFT; ws.merge_cells("B6:E6"); c6b = ws['B6']; c6b.value = product_name; c6b.font = FONT_14; c6b.alignment = ALIGN_LEFT
        month_groups = []
        for i in range(eff_days):
            d = start_dt + timedelta(days=i); m_key = (d.year, d.month)
            if not month_groups or month_groups[-1][0] != m_key: month_groups.append([m_key, i, i]) 
            else: month_groups[-1][2] = i
        for m_key, s_idx, e_idx in month_groups:
            start_col = 6 + s_idx; end_col = 6 + e_idx
            ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col); c = ws.cell(6, start_col); c.value = f"{m_key[1]}月"; c.font = FONT_BOLD; c.alignment = ALIGN_LEFT; c.border = BORDER_ALL_MEDIUM
        for c_idx in range(1, total_cols + 1):
            c = ws.cell(6, c_idx); t, b, l, r = BS_MEDIUM, BS_MEDIUM, None, None
            if c_idx == 1: l = BS_MEDIUM 
            if c_idx == total_cols: r = BS_MEDIUM 
            if c_idx == 6: l = None 
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l) if l else None, right=Side(style=r) if r else None)
        draw_outer_border_fast(ws, 6, 6, 1, 5); ws.cell(6, 5).border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_MEDIUM), right=Side(style=None))
        
        header_start_row = 7; headers = ["頻道", "播出地區", "播出店數", "播出時間", "秒數\n規格"]
        for i, h in enumerate(headers):
            c_idx = i + 1; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_16_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == 1: l = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).font = FONT_16

        curr = start_dt
        for i in range(eff_days):
            col_idx = 6 + i; c7 = ws.cell(header_start_row, col_idx); c7.value = curr.day; c7.font = FONT_16_BOLD; c7.alignment = ALIGN_CENTER; c7.border = BORDER_ALL_MEDIUM
            c7.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            c8 = ws.cell(header_start_row+1, col_idx); c8.value = ["日","一","二","三","四","五","六"][(curr.weekday()+1)%7]; c8.font = FONT_16; c8.alignment = ALIGN_CENTER
            style_left = BS_MEDIUM if col_idx == 6 else BS_THIN
            c8.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=style_left), right=Side(style=BS_THIN)); 
            if curr.weekday() >= 5: c8.fill = FILL_WEEKEND
            curr += timedelta(days=1)

        end_headers = ["檔次", "定價", "專案價"]; 
        for i, h in enumerate(end_headers):
            c_idx = end_c_start + i; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_16_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == total_cols: r = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).font = FONT_16

        date_start_col = 6
        for c_idx in range(date_start_col, total_cols + 1):
            c7 = ws.cell(header_start_row, c_idx); c7.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            if c_idx == date_start_col: set_border(c7, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c7, right=BS_MEDIUM)
            c8 = ws.cell(8, c_idx); c8.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN)); c8.font = FONT_16
            if c_idx == date_start_col: set_border(c8, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c8, right=BS_MEDIUM)
        ws.cell(header_start_row + 1, end_c_start - 1).border = Border(top=SIDE_THIN, bottom=SIDE_THIN, left=SIDE_THIN, right=SIDE_MEDIUM)

        curr_row = header_start_row + 2; grouped_data = {"全家廣播": sorted([r for r in rows if r["media"]=="全家廣播"], key=lambda x:x['seconds']), "新鮮視": sorted([r for r in rows if r["media"]=="新鮮視"], key=lambda x:x['seconds']), "家樂福": sorted([r for r in rows if r["media"]=="家樂福"], key=lambda x:x['seconds'])}
        total_store_count = 0; total_list_sum = 0

        for m_key, data in grouped_data.items():
            if not data: continue
            start_merge = curr_row; d_name = f"全家便利商店\n{m_key}廣告" if m_key != "家樂福" else "家樂福"
            for idx, r in enumerate(data):
                ws.row_dimensions[curr_row].height = 54; ws.cell(curr_row, 1, d_name).alignment = ALIGN_CENTER; ws.cell(curr_row, 2, r['region']).alignment = ALIGN_CENTER
                p_num = int(r.get('program_num', 0)); total_store_count += p_num; suffix = "面" if m_key == "新鮮視" else "店"; ws.cell(curr_row, 3, f"{p_num:,}{suffix}").alignment = ALIGN_CENTER
                ws.cell(curr_row, 4, r['daypart']).alignment = ALIGN_CENTER
                sec = r['seconds']; sec_txt = f"{sec}秒\n影片/影像 1920x1080 (mp4)" if m_key == "新鮮視" else f"{sec}秒廣告"; c_spec = ws.cell(curr_row, 5, sec_txt); c_spec.alignment = ALIGN_CENTER; c_spec.font = Font(name=FONT_MAIN, size=10)
                row_sum = 0
                for d_idx in range(eff_days):
                    if d_idx < len(r['schedule']): val = r['schedule'][d_idx]; v = val if isinstance(val, (int, float)) else 0; row_sum += v; c = ws.cell(curr_row, 6+d_idx); c.value = "" if (val == 0 or val is None) else val; c.alignment = ALIGN_CENTER; c.font = FONT_WEEKEND if (start_dt + timedelta(days=d_idx)).weekday() >= 5 else FONT_DAILY; c.border = BORDER_ALL_THIN
                ws.cell(curr_row, end_c_start, row_sum).alignment = ALIGN_CENTER
                rate_val = r['rate_display']; 
                if isinstance(rate_val, (int, float)): total_list_sum += rate_val
                ws.cell(curr_row, end_c_start+1, rate_val).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER 
                pkg = r['pkg_display']; 
                if r.get('is_pkg_member'): pkg = r['nat_pkg_display'] if idx == 0 else None
                elif r.get('is_rebate'): pkg = r.get('pkg_display', '回饋') if (idx == 0 or not data[idx-1].get('is_rebate') or data[idx-1].get('is_bonus_rebate') != r.get('is_bonus_rebate')) else None
                elif r.get('is_custom_bonus'): pkg = r.get('pkg_display', '加贈') if (idx == 0 or not data[idx-1].get('is_custom_bonus')) else None
                if pkg is not None: ws.cell(curr_row, end_c_start+2, pkg).alignment = ALIGN_CENTER; ws.cell(curr_row, end_c_start+2).number_format = FMT_MONEY if isinstance(pkg, (int, float)) else '@'
                for c_idx in range(1, total_cols + 1):
                    c = ws.cell(curr_row, c_idx); c.border = BORDER_ALL_THIN
                    if c_idx < 6 or c_idx >= end_c_start: c.font = FONT_16
                set_border(ws.cell(curr_row, 5), right=BS_MEDIUM); curr_row += 1
            ws.merge_cells(start_row=start_merge, start_column=1, end_row=curr_row-1, end_column=1)
            i = 0
            while i < len(data):
                if data[i].get('is_pkg_member'):
                    j = i
                    while j < len(data) and data[j].get('is_pkg_member'): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=end_c_start+2, end_row=start_merge+j-1, end_column=end_c_start+2)
                    i = j
                elif data[i].get('is_rebate'):
                    j = i
                    while j < len(data) and data[j].get('is_rebate') and data[j].get('is_bonus_rebate') == data[i].get('is_bonus_rebate'): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=end_c_start+2, end_row=start_merge+j-1, end_column=end_c_start+2)
                    i = j
                elif data[i].get('is_custom_bonus'):
                    j = i
                    while j < len(data) and data[j].get('is_custom_bonus'): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=end_c_start+2, end_row=start_merge+j-1, end_column=end_c_start+2)
                    i = j
                else:
                    i += 1
            draw_outer_border_fast(ws, start_merge, curr_row-1, 1, total_cols)

        ws.row_dimensions[curr_row].height = 54; ws.cell(curr_row, 3, total_store_count).number_format = FMT_NUMBER; ws.cell(curr_row, 3).alignment = ALIGN_CENTER; ws.cell(curr_row, 3).font = FONT_16_BOLD
        ws.cell(curr_row, 5, "Total").alignment = ALIGN_CENTER; ws.cell(curr_row, 5).font = FONT_16_BOLD
        for d_idx in range(eff_days): daily_sum = sum([r['schedule'][d_idx] if d_idx < len(r['schedule']) and isinstance(r['schedule'][d_idx], (int, float)) else 0 for r in rows]); c = ws.cell(curr_row, 6+d_idx); c.value = "" if daily_sum == 0 else daily_sum; c.alignment = ALIGN_CENTER; c.font = FONT_WEEKEND if (start_dt + timedelta(days=d_idx)).weekday() >= 5 else FONT_DAILY
        ws.cell(curr_row, end_c_start, sum([sum(r['schedule']) for r in rows])).alignment = ALIGN_CENTER; ws.cell(curr_row, end_c_start).font = FONT_16_BOLD
        ws.cell(curr_row, end_c_start+1, total_list_sum).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).font = FONT_16_BOLD; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER
        pkg_total_val = (pkg_total if pkg_total is not None else budget)
        ws.cell(curr_row, end_c_start+2, pkg_total_val).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+2).font = FONT_16_BOLD; ws.cell(curr_row, end_c_start+2).alignment = ALIGN_CENTER
        for c_idx in range(1, total_cols+1): ws.cell(curr_row, c_idx).border = BORDER_ALL_THIN
        draw_outer_border_fast(ws, curr_row, curr_row, 1, total_cols)
        for c_idx in range(1, total_cols+1): set_border(ws.cell(curr_row, c_idx), bottom=BS_MEDIUM)
        set_border(ws.cell(curr_row, 5), right=BS_MEDIUM)
        ws.cell(curr_row, end_c_start - 1).border = Border(top=SIDE_MEDIUM, bottom=SIDE_MEDIUM, left=SIDE_THIN, right=SIDE_MEDIUM)
        curr_row += 1

        vat = _round_half_up(budget * 0.05); grand_total = budget + vat
        footer_stack = [("製作", prod), ("5% VAT", vat), ("Grand Total", grand_total)]
        for lbl, val in footer_stack:
            ws.row_dimensions[curr_row].height = 30; c_l = ws.cell(curr_row, end_c_start+1); c_l.value = lbl; c_l.alignment = ALIGN_RIGHT; c_l.font = FONT_16
            c_v = ws.cell(curr_row, end_c_start+2); c_v.value = val; c_v.number_format = FMT_MONEY; c_v.alignment = ALIGN_CENTER; c_v.font = FONT_16 
            t, b, l, r = BS_THIN, BS_THIN, BS_MEDIUM, BS_THIN; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_l.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            t, b, l, r = BS_THIN, BS_THIN, BS_THIN, BS_MEDIUM; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_v.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            curr_row += 1
        
        # Remarks 欄位起點比照鉑霖：<14 天對齊秒數規格欄，>=14 天對齊右側欄
        curr_row += 1; start_footer = curr_row; r_col_start = 5 if eff_days < 14 else 6
        ws.row_dimensions[start_footer].height = 25; ws.cell(start_footer, r_col_start).value = "Remarks：本排程表經雙方確認後視同合約之延伸，具同等法律約束力與效力"
        ws.cell(start_footer, r_col_start).font = Font(name=FONT_MAIN, size=18, bold=True)
        def _remark_chars_per_line(start_col, end_col):
            width_sum = 0.0
            for cidx in range(start_col, end_col + 1):
                letter = get_column_letter(cidx)
                w = ws.column_dimensions[letter].width
                width_sum += float(w if w is not None else 8.43)
            return max(36, int(width_sum * 0.78))

        def _char_visual_width(ch):
            if ch == "\t":
                return 2.0
            if ch.isspace():
                return 0.7
            return 2.0 if unicodedata.east_asian_width(ch) in ("W", "F") else 1.0

        def _simulate_wrapped_lines(text, max_units):
            t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
            t = "\n".join([seg.strip() for seg in t.split("\n") if seg.strip()])
            if not t:
                return [""]
            effective_units = max(24.0, float(max_units) * 0.88)
            result = []
            punct = set("。；，,、 ")
            for raw_seg in t.split("\n"):
                seg = raw_seg.strip()
                if not seg:
                    continue
                line = ""
                used = 0.0
                last_break_idx = -1
                i = 0
                while i < len(seg):
                    ch = seg[i]
                    w = _char_visual_width(ch)
                    if used + w <= effective_units or not line:
                        line += ch
                        used += w
                        if ch in punct:
                            last_break_idx = len(line)
                        i += 1
                        continue
                    if last_break_idx > 0:
                        result.append(line[:last_break_idx].rstrip())
                        line = line[last_break_idx:].lstrip()
                        used = sum(_char_visual_width(c) for c in line)
                        last_break_idx = -1
                    else:
                        result.append(line.rstrip())
                        line = ""
                        used = 0.0
                if line.strip():
                    result.append(line.rstrip())
            return result or [""]

        r_row = start_footer
        for rm in remarks_list:
            is_red = rm.strip().startswith("1.") or rm.strip().startswith("4.")
            is_blue = rm.strip().startswith("6.")
            color = "FF0000" if is_red else ("0000FF" if is_blue else "000000")
            max_units = _remark_chars_per_line(r_col_start, total_cols)
            lines = _simulate_wrapped_lines(rm, max_units=max_units)
            # 每個模擬行獨立輸出成一列，避免 PDF 二次換行造成重疊或截字
            for line_text in lines:
                try:
                    ws.merge_cells(start_row=r_row + 1, start_column=r_col_start, end_row=r_row + 1, end_column=total_cols)
                except Exception:
                    pass
                r_row += 1
                if len(lines) == 1:
                    ws.row_dimensions[r_row].height = 28 if eff_days < 14 else 30
                else:
                    ws.row_dimensions[r_row].height = 24 if eff_days < 14 else 26
                c = ws.cell(r_row, r_col_start)
                c.value = line_text
                c.font = Font(name=FONT_MAIN, size=(16 if eff_days < 14 else 18), color=color)
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

        sig_col_start = 1
        for _r in (start_footer, start_footer+1, start_footer+2, start_footer+3): ws.row_dimensions[_r].height = 28
        ws.cell(start_footer, sig_col_start).value = "乙         方："; ws.cell(start_footer, sig_col_start).font = Font(name=FONT_MAIN, size=20)
        ws.cell(start_footer+1, sig_col_start+1).value = client_name; ws.cell(start_footer+1, sig_col_start+1).font = Font(name=FONT_MAIN, size=20)
        ws.cell(start_footer+2, sig_col_start).value = "統一編號："; ws.cell(start_footer+2, sig_col_start).font = Font(name=FONT_MAIN, size=20)
        
        # 填入 Excel 統編 (聲活格式) - 改為B欄
        ws.cell(start_footer+2, sig_col_start+1).value = tax_id
        ws.cell(start_footer+2, sig_col_start+1).font = Font(name=FONT_MAIN, size=20)
        
        ws.cell(start_footer+3, sig_col_start).value = "客戶簽章："; ws.cell(start_footer+3, sig_col_start).font = Font(name=FONT_MAIN, size=20)

        return r_row + 2

    # ---------------------------------------------------------
    # Sub-Engine: Bolin (鉑霖格式 - 已還原舊版樣式 + 統編對齊修正)
    # ---------------------------------------------------------
    def render_bolin_optimized(ws, start_dt, end_dt, rows, budget, prod, pkg_total=None):
        SIDE_DOUBLE = Side(style='double')
        logo_bytes = get_cloud_logo_bytes()
        eff_days = (end_dt - start_dt).days + 1
        end_c_start = 6 + eff_days
        total_cols = end_c_start + 2
        
        # 欄寬設定 (舊版配置)
        ws.column_dimensions['A'].width = 21.0
        ws.column_dimensions['B'].width = 21.0
        ws.column_dimensions['C'].width = 13.8
        ws.column_dimensions['D'].width = 19.4
        ws.column_dimensions['E'].width = 27.0
        for i in range(eff_days): 
            ws.column_dimensions[get_column_letter(6 + i)].width = 8.1
        ws.column_dimensions[get_column_letter(end_c_start)].width = 9.5
        ws.column_dimensions[get_column_letter(end_c_start+1)].width = 36.0
        # 專案價欄位加寬到比 logo 更寬，避免 logo 視覺外溢
        ws.column_dimensions[get_column_letter(end_c_start+2)].width = 34.0
        
        ROW_H_MAP = {1:70, 2:33.5, 3:33.5, 4:46, 5:40, 6:35, 7:35}
        for r, h in ROW_H_MAP.items(): ws.row_dimensions[r].height = h
        
        # 標題與 Logo (舊版樣式)
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1"); c1 = ws['A1']
        c1.value = "鉑霖行動行銷-媒體計劃排程表 Mobi Media Schedule"
        c1.font = Font(name=FONT_MAIN, size=28, bold=True); c1.alignment = ALIGN_LEFT 
        if logo_bytes:
            try: 
                img = OpenpyxlImage(io.BytesIO(logo_bytes))
                scale = 120 / img.height
                img.height = 120
                img.width = int(img.width * scale)
                # Logo 改錨在「專案價」欄
                col_letter = get_column_letter(total_cols)
                img.anchor = f"{col_letter}1"
                ws.add_image(img)
            except Exception: pass

        # TO / FROM 區塊 (舊版樣式)
        c2a = ws['A2']; c2a.value = "TO："; c2a.font = Font(name=FONT_MAIN, size=20, bold=True, color="FF0000"); c2a.alignment = ALIGN_LEFT
        ws.merge_cells(f"B2:{get_column_letter(total_cols)}2"); c2b = ws['B2']; c2b.value = client_name
        c2b.font = Font(name=FONT_MAIN, size=20, bold=True, color="FF0000"); c2b.alignment = ALIGN_LEFT
        
        c3a = ws['A3']; c3a.value = "FROM："; c3a.font = Font(name=FONT_MAIN, size=20, bold=True); c3a.alignment = ALIGN_LEFT
        ws.merge_cells(f"B3:{get_column_letter(total_cols)}3"); c3b = ws['B3']; c3b.value = f"鉑霖行動行銷 {sales_person}"
        c3b.font = Font(name=FONT_MAIN, size=20, bold=True); c3b.alignment = ALIGN_LEFT

        # 客戶資料與規格 (舊版樣式)
        unique_secs = sorted(list(set([r['seconds'] for r in rows])))
        sec_str = " ".join([f"{s}秒廣告" for s in unique_secs])
        period_str = f"執行期間：{start_dt.strftime('%Y.%m.%d')} - {end_dt.strftime('%Y.%m.%d')}"
        
        c4a = ws['A4']; c4a.value = "客戶名稱："; c4a.font = Font(name=FONT_MAIN, size=20, bold=True); c4a.alignment = ALIGN_LEFT
        ws.merge_cells("B4:E4"); c4b = ws['B4']; c4b.value = client_name; c4b.font = Font(name=FONT_MAIN, size=20, bold=True); c4b.alignment = ALIGN_LEFT
        
        spec_merge_start = "F4"; spec_merge_end = f"{get_column_letter(end_c_start)}4"
        ws.merge_cells(f"{spec_merge_start}:{spec_merge_end}"); c4f = ws['F4']; c4f.value = f"廣告規格：{sec_str}"
        c4f.font = Font(name=FONT_MAIN, size=20, bold=True); c4f.alignment = ALIGN_LEFT
        
        # 執行期間放在定價區，靠左單行展開（跨到專案價欄）
        ws.merge_cells(f"{get_column_letter(end_c_start+1)}4:{get_column_letter(total_cols)}4")
        c4_r = ws[f"{get_column_letter(end_c_start+1)}4"]; c4_r.value = period_str
        c4_r.font = Font(name=FONT_MAIN, size=20, bold=True)
        c4_r.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=False)
        draw_outer_border_fast(ws, 4, 4, 1, total_cols)

        # 產品名稱與月份 (舊版樣式)
        c5a = ws['A5']; c5a.value = "廣告名稱："; c5a.font = Font(name=FONT_MAIN, size=20, bold=True); c5a.alignment = ALIGN_LEFT
        ws.merge_cells("B5:E5"); c5b = ws['B5']; c5b.value = product_name; c5b.font = Font(name=FONT_MAIN, size=20, bold=True); c5b.alignment = ALIGN_LEFT
        
        month_groups = []
        for i in range(eff_days):
            d = start_dt + timedelta(days=i); m_key = (d.year, d.month)
            if not month_groups or month_groups[-1][0] != m_key: month_groups.append([m_key, i, i]) 
            else: month_groups[-1][2] = i
        for m_key, s_idx, e_idx in month_groups:
            start_col = 6 + s_idx; end_col = 6 + e_idx
            ws.merge_cells(start_row=5, start_column=start_col, end_row=5, end_column=end_col)
            c = ws.cell(5, start_col); c.value = f"{m_key[1]}月"; c.font = FONT_BOLD; c.alignment = ALIGN_LEFT 
            
        for c_idx in range(1, total_cols + 1):
            c = ws.cell(5, c_idx); t, b, l, r = BS_MEDIUM, BS_MEDIUM, None, None
            if c_idx == 1: l = BS_MEDIUM 
            if c_idx == total_cols: r = BS_MEDIUM 
            if c_idx == 6: l = None 
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l) if l else None, right=Side(style=r) if r else None)
        draw_outer_border_fast(ws, 5, 5, 1, 5); ws.cell(5, 5).border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_MEDIUM), right=Side(style=None))

        # 表頭 (舊版樣式)
        header_start_row = 6; headers = ["頻道", "播出地區", "播出店數", "播出時間", "秒數\n規格"]
        for i, h in enumerate(headers):
            c_idx = i + 1; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_16_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == 1: l = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).font = FONT_16

        curr = start_dt
        for i in range(eff_days):
            col_idx = 6 + i; c6 = ws.cell(header_start_row, col_idx); c6.value = curr.day; c6.font = FONT_16_BOLD; c6.alignment = ALIGN_CENTER; c6.border = BORDER_ALL_MEDIUM; c6.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            c7 = ws.cell(header_start_row+1, col_idx); c7.value = ["日","一","二","三","四","五","六"][(curr.weekday()+1)%7]; c7.font = FONT_16; c7.alignment = ALIGN_CENTER; style_left = BS_MEDIUM if col_idx == 6 else BS_THIN; c7.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=style_left), right=Side(style=BS_THIN))
            if curr.weekday() >= 5: c7.fill = FILL_WEEKEND
            curr += timedelta(days=1)

        end_headers = ["檔次", "定價", "專案價"]; 
        for i, h in enumerate(end_headers):
            c_idx = end_c_start + i; ws.merge_cells(start_row=header_start_row, start_column=c_idx, end_row=header_start_row+1, end_column=c_idx); c = ws.cell(header_start_row, c_idx); c.value = h; c.font = FONT_16_BOLD; c.alignment = ALIGN_CENTER
            t, b, l, r = BS_MEDIUM, BS_THIN, BS_THIN, BS_THIN; 
            if c_idx == total_cols: r = BS_MEDIUM
            c.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=l), right=Side(style=r)); ws.cell(header_start_row+1, c_idx).font = FONT_16

        date_start_col = 6
        for c_idx in range(date_start_col, total_cols + 1):
            c7 = ws.cell(header_start_row, c_idx); c7.border = Border(top=Side(style=BS_MEDIUM), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN))
            if c_idx == date_start_col: set_border(c7, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c7, right=BS_MEDIUM)
            c8 = ws.cell(8, c_idx); c8.border = Border(top=Side(style=BS_THIN), bottom=Side(style=BS_THIN), left=Side(style=BS_THIN), right=Side(style=BS_THIN)); c8.font = FONT_16
            if c_idx == date_start_col: set_border(c8, left=BS_MEDIUM)
            if c_idx == total_cols: set_border(c8, right=BS_MEDIUM)
        ws.cell(header_start_row + 1, end_c_start - 1).border = Border(top=SIDE_THIN, bottom=SIDE_THIN, left=SIDE_THIN, right=SIDE_MEDIUM)

        # 內容資料 (共用邏輯)
        curr_row = header_start_row + 2; grouped_data = {"全家廣播": sorted([r for r in rows if r["media"]=="全家廣播"], key=lambda x:x['seconds']), "新鮮視": sorted([r for r in rows if r["media"]=="新鮮視"], key=lambda x:x['seconds']), "家樂福": sorted([r for r in rows if r["media"]=="家樂福"], key=lambda x:x['seconds'])}
        total_store_count = 0; total_list_sum = 0
        for m_key, data in grouped_data.items():
            if not data: continue
            start_merge = curr_row; d_name = f"全家便利商店\n{m_key}廣告" if m_key != "家樂福" else "家樂福"
            for idx, r in enumerate(data):
                ws.row_dimensions[curr_row].height = 54; ws.cell(curr_row, 1, d_name).alignment = ALIGN_CENTER; ws.cell(curr_row, 2, r['region']).alignment = ALIGN_CENTER
                p_num = int(r.get('program_num', 0)); total_store_count += p_num; suffix = "面" if m_key == "新鮮視" else "店"; ws.cell(curr_row, 3, f"{p_num:,}{suffix}").alignment = ALIGN_CENTER
                ws.cell(curr_row, 4, r['daypart']).alignment = ALIGN_CENTER
                sec = r['seconds']; sec_txt = f"{sec}秒\n影片/影像 1920x1080 (mp4)" if m_key == "新鮮視" else f"{sec}秒廣告"; c_spec = ws.cell(curr_row, 5, sec_txt); c_spec.alignment = ALIGN_CENTER; c_spec.font = Font(name=FONT_MAIN, size=10)
                row_sum = 0
                for d_idx in range(eff_days):
                    if d_idx < len(r['schedule']): val = r['schedule'][d_idx]; v = val if isinstance(val, (int, float)) else 0; row_sum += v; c = ws.cell(curr_row, 6+d_idx); c.value = "" if (val == 0 or val is None) else val; c.alignment = ALIGN_CENTER; c.font = FONT_WEEKEND if (start_dt + timedelta(days=d_idx)).weekday() >= 5 else FONT_DAILY; c.border = BORDER_ALL_THIN
                ws.cell(curr_row, end_c_start, row_sum).alignment = ALIGN_CENTER
                rate_val = r['rate_display']; 
                if isinstance(rate_val, (int, float)): total_list_sum += rate_val
                ws.cell(curr_row, end_c_start+1, rate_val).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER 
                pkg = r['pkg_display']; 
                if r.get('is_pkg_member'): pkg = r['nat_pkg_display'] if idx == 0 else None
                elif r.get('is_rebate'): pkg = r.get('pkg_display', '回饋') if (idx == 0 or not data[idx-1].get('is_rebate') or data[idx-1].get('is_bonus_rebate') != r.get('is_bonus_rebate')) else None
                elif r.get('is_custom_bonus'): pkg = r.get('pkg_display', '加贈') if (idx == 0 or not data[idx-1].get('is_custom_bonus')) else None
                if pkg is not None: ws.cell(curr_row, end_c_start+2, pkg).alignment = ALIGN_CENTER; ws.cell(curr_row, end_c_start+2).number_format = FMT_MONEY if isinstance(pkg, (int, float)) else '@'
                for c_idx in range(1, total_cols + 1):
                    c = ws.cell(curr_row, c_idx); c.border = BORDER_ALL_THIN
                    if c_idx < 6 or c_idx >= end_c_start: c.font = FONT_16
                set_border(ws.cell(curr_row, 5), right=BS_MEDIUM); curr_row += 1
            ws.merge_cells(start_row=start_merge, start_column=1, end_row=curr_row-1, end_column=1)
            i = 0
            while i < len(data):
                if data[i].get('is_pkg_member'):
                    j = i
                    while j < len(data) and data[j].get('is_pkg_member'): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=end_c_start+2, end_row=start_merge+j-1, end_column=end_c_start+2)
                    i = j
                elif data[i].get('is_rebate'):
                    j = i
                    while j < len(data) and data[j].get('is_rebate') and data[j].get('is_bonus_rebate') == data[i].get('is_bonus_rebate'): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=end_c_start+2, end_row=start_merge+j-1, end_column=end_c_start+2)
                    i = j
                elif data[i].get('is_custom_bonus'):
                    j = i
                    while j < len(data) and data[j].get('is_custom_bonus'): j += 1
                    if j > i + 1: ws.merge_cells(start_row=start_merge+i, start_column=end_c_start+2, end_row=start_merge+j-1, end_column=end_c_start+2)
                    i = j
                else:
                    i += 1
            draw_outer_border_fast(ws, start_merge, curr_row-1, 1, total_cols)

        # 總計行 (共用邏輯)
        ws.row_dimensions[curr_row].height = 54; ws.cell(curr_row, 3, total_store_count).number_format = FMT_NUMBER; ws.cell(curr_row, 3).alignment = ALIGN_CENTER; ws.cell(curr_row, 3).font = FONT_16_BOLD
        ws.cell(curr_row, 5, "Total").alignment = ALIGN_CENTER; ws.cell(curr_row, 5).font = FONT_16_BOLD
        for d_idx in range(eff_days): daily_sum = sum([r['schedule'][d_idx] if d_idx < len(r['schedule']) and isinstance(r['schedule'][d_idx], (int, float)) else 0 for r in rows]); c = ws.cell(curr_row, 6+d_idx); c.value = "" if daily_sum == 0 else daily_sum; c.alignment = ALIGN_CENTER; c.font = FONT_WEEKEND if (start_dt + timedelta(days=d_idx)).weekday() >= 5 else FONT_DAILY
        ws.cell(curr_row, end_c_start, sum([sum(r['schedule']) for r in rows])).alignment = ALIGN_CENTER; ws.cell(curr_row, end_c_start).font = FONT_16_BOLD
        ws.cell(curr_row, end_c_start+1, total_list_sum).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+1).font = FONT_16_BOLD; ws.cell(curr_row, end_c_start+1).alignment = ALIGN_CENTER
        pkg_total_val = (pkg_total if pkg_total is not None else budget)
        ws.cell(curr_row, end_c_start+2, pkg_total_val).number_format = FMT_MONEY; ws.cell(curr_row, end_c_start+2).font = FONT_16_BOLD; ws.cell(curr_row, end_c_start+2).alignment = ALIGN_CENTER
        for c_idx in range(1, total_cols+1): ws.cell(curr_row, c_idx).border = BORDER_ALL_THIN
        draw_outer_border_fast(ws, curr_row, curr_row, 1, total_cols)
        for c_idx in range(1, total_cols+1): set_border(ws.cell(curr_row, c_idx), bottom=BS_MEDIUM)
        set_border(ws.cell(curr_row, 5), right=BS_MEDIUM)
        ws.cell(curr_row, end_c_start - 1).border = Border(top=SIDE_MEDIUM, bottom=SIDE_MEDIUM, left=SIDE_THIN, right=SIDE_MEDIUM)
        curr_row += 1

        vat = _round_half_up(budget * 0.05); grand_total = budget + vat
        footer_stack = [("製作", prod), ("5% VAT", vat), ("Grand Total", grand_total)]
        for lbl, val in footer_stack:
            ws.row_dimensions[curr_row].height = 30; c_l = ws.cell(curr_row, end_c_start+1); c_l.value = lbl; c_l.alignment = ALIGN_RIGHT; c_l.font = FONT_16
            c_v = ws.cell(curr_row, end_c_start+2); c_v.value = val; c_v.number_format = FMT_MONEY; c_v.alignment = ALIGN_CENTER; c_v.font = FONT_16 
            t, b, l, r = BS_THIN, BS_THIN, BS_MEDIUM, BS_THIN; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_l.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            t, b, l, r = BS_THIN, BS_THIN, BS_THIN, BS_MEDIUM; 
            if lbl == "Grand Total": b = BS_MEDIUM 
            c_v.border = Border(top=Side(style=t), bottom=Side(style=b), left=Side(style=l), right=Side(style=r))
            curr_row += 1
        
        # Footer & 簽名區 (舊版樣式 + 統編對齊修正)
        # Remarks 欄位起點：<14 天對齊「秒數規格」欄；>=14 天對齊其右側欄
        curr_row += 1; start_footer = curr_row; r_col_start = 5 if eff_days < 14 else 6
        ws.row_dimensions[start_footer].height = 25; ws.cell(start_footer, r_col_start).value = "Remarks：本排程表經雙方確認後視同合約之延伸，具同等法律約束力與效力"
        ws.cell(start_footer, r_col_start).font = Font(name=FONT_MAIN, size=18, bold=True)
        def _remark_chars_per_line(start_col, end_col):
            """依 remarks 合併區實際欄寬估算每列可容納的視覺寬度單位。"""
            width_sum = 0.0
            for cidx in range(start_col, end_col + 1):
                letter = get_column_letter(cidx)
                w = ws.column_dimensions[letter].width
                width_sum += float(w if w is not None else 8.43)
            # 保守估算：PDF 轉檔常比 Excel 預覽更容易提早換行
            return max(36, int(width_sum * 0.78))

        def _char_visual_width(ch):
            if ch == "\t":
                return 2.0
            if ch.isspace():
                return 0.7
            # CJK 全形字元視覺寬度較大
            return 2.0 if unicodedata.east_asian_width(ch) in ("W", "F") else 1.0

        def _simulate_wrapped_lines(text, max_units):
            """
            先模擬最終視覺斷行（可延伸就延伸、超出才換行），
            再用此結果回填儲存格與列高，避免靠死板固定字數。
            """
            t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
            t = "\n".join([seg.strip() for seg in t.split("\n") if seg.strip()])
            if not t:
                return [""]
            effective_units = max(24.0, float(max_units) * 0.88)
            result = []
            punct = set("。；，,、 ")
            for raw_seg in t.split("\n"):
                seg = raw_seg.strip()
                if not seg:
                    continue
                line = ""
                used = 0.0
                last_break_idx = -1
                last_break_used = 0.0
                i = 0
                while i < len(seg):
                    ch = seg[i]
                    w = _char_visual_width(ch)
                    if used + w <= effective_units or not line:
                        line += ch
                        used += w
                        if ch in punct:
                            last_break_idx = len(line)
                            last_break_used = used
                        i += 1
                        continue
                    if last_break_idx > 0:
                        result.append(line[:last_break_idx].rstrip())
                        remain = line[last_break_idx:].lstrip()
                        line = remain
                        used = sum(_char_visual_width(c) for c in line)
                        last_break_idx = -1
                        last_break_used = 0.0
                    else:
                        result.append(line.rstrip())
                        line = ""
                        used = 0.0
                        last_break_idx = -1
                        last_break_used = 0.0
                if line.strip():
                    result.append(line.rstrip())
            return result or [""]

        r_row = start_footer
        for rm in remarks_list:
            is_red = rm.strip().startswith("1.") or rm.strip().startswith("4.")
            is_blue = rm.strip().startswith("6.")
            color = "FF0000" if is_red else ("0000FF" if is_blue else "000000")

            max_units = _remark_chars_per_line(r_col_start, total_cols)
            lines = _simulate_wrapped_lines(rm, max_units=max_units)
            # 以「每個模擬行 = 一個實際列」輸出，避免 PDF 端再次自動換行造成重疊
            for idx, line_text in enumerate(lines):
                try:
                    ws.merge_cells(
                        start_row=r_row + 1,
                        start_column=r_col_start,
                        end_row=r_row + 1,
                        end_column=total_cols,
                    )
                except Exception:
                    pass
                r_row += 1
                if len(lines) == 1:
                    ws.row_dimensions[r_row].height = 28 if eff_days <= 14 else 30
                else:
                    ws.row_dimensions[r_row].height = 24 if eff_days <= 14 else 26
                c = ws.cell(r_row, r_col_start)
                c.value = line_text
                c.font = Font(name=FONT_MAIN, size=(16 if eff_days <= 14 else 18), color=color)
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)

        sig_col_start = 1
        
        # 乙方區塊上方分隔線：延伸為整張表寬（與上方標題分隔線一致）
        for c_idx in range(1, total_cols + 1):
            set_border(ws.cell(start_footer, c_idx), top=BS_MEDIUM)
        
        # 乙方
        ws.cell(start_footer, sig_col_start).value = "乙         方："; ws.cell(start_footer, sig_col_start).font = Font(name=FONT_MAIN, size=20)
        # 客戶名稱 (B欄)
        ws.cell(start_footer+1, sig_col_start+1).value = client_name; ws.cell(start_footer+1, sig_col_start+1).font = Font(name=FONT_MAIN, size=20)
        
        # 統編 (修改處：值移到 B 欄)
        ws.cell(start_footer+2, sig_col_start).value = "統一編號："; ws.cell(start_footer+2, sig_col_start).font = Font(name=FONT_MAIN, size=20)
        ws.cell(start_footer+2, sig_col_start+1).value = tax_id; ws.cell(start_footer+2, sig_col_start+1).font = Font(name=FONT_MAIN, size=20)
        
        # 客戶簽章
        ws.cell(start_footer+3, sig_col_start).value = "客戶簽章："; ws.cell(start_footer+3, sig_col_start).font = Font(name=FONT_MAIN, size=20)

        return r_row + 2
        
    # Main Execution of Excel Generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True

    month_ranges = split_period_by_months(start_dt, end_dt)
    total_days = (end_dt - start_dt).days + 1

    # 走期 31 天內（含跨月）不拆表，一張表呈現；超過 31 天且跨月才拆成多頁／多工作表
    if len(month_ranges) > 1 and total_days > 31:
        # 走期超過一個月且超過 31 天：多工作表／多頁，金額依該月天數比例攤分
        def build_rows_and_budget(m_start, m_end):
            try:
                day_offset = (m_start - start_dt).days
            except Exception:
                day_offset = 0
            day_offset = max(0, day_offset)
            days_in_month = (m_end - m_start).days + 1
            ratio = days_in_month / total_days if total_days else 0
            rows_month = []
            for r in rows:
                r2 = dict(r)
                sch = r.get("schedule", [])
                r2["schedule"] = sch[day_offset : day_offset + days_in_month]
                if isinstance(r.get("rate_display"), (int, float)):
                    r2["rate_display"] = int(round(r["rate_display"] * ratio))
                if isinstance(r.get("pkg_display"), (int, float)):
                    r2["pkg_display"] = int(round(r["pkg_display"] * ratio))
                if isinstance(r.get("nat_pkg_display"), (int, float)):
                    r2["nat_pkg_display"] = int(round(r["nat_pkg_display"] * ratio))
                rows_month.append(r2)
            budget_month = int(round(final_budget_val * ratio))
            prod_month = int(round(prod_cost * ratio))
            return rows_month, budget_month, prod_month

        if format_type == "東吳":
            # 東吳：每月獨立工作表，每頁完整標題與備註（與 HTML 兩頁呈現一致）
            for i, (m_start, m_end) in enumerate(month_ranges):
                rows_month, budget_month, prod_month = build_rows_and_budget(m_start, m_end)
                if i == 0:
                    ws = wb.active
                    ws.title = f"{m_start.month}月"
                else:
                    ws = wb.create_sheet(title=f"{m_start.month}月")
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToPage = True
                render_dongwu_optimized(ws, m_start, m_end, rows_month, budget_month, prod_month, start_row=1, skip_footer=False, first_block=True, skip_title=False)
        else:
            # 聲活/鉑霖：每月一個工作表
            for i, (m_start, m_end) in enumerate(month_ranges):
                rows_month, budget_month, prod_month = build_rows_and_budget(m_start, m_end)
                if i == 0:
                    ws = wb.active
                    ws.title = f"{m_start.month}月"
                else:
                    ws = wb.create_sheet(title=f"{m_start.month}月")
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToPage = True
                if format_type == "聲活":
                    render_shenghuo_optimized(ws, m_start, m_end, rows_month, budget_month, prod_month)
                else:
                    render_bolin_optimized(ws, m_start, m_end, rows_month, budget_month, prod_month)
    else:
        # 單月：維持原本單一工作表
        if format_type == "東吳":
            render_dongwu_optimized(ws, start_dt, end_dt, rows, final_budget_val, prod_cost)
        elif format_type == "聲活":
            render_shenghuo_optimized(ws, start_dt, end_dt, rows, final_budget_val, prod_cost)
        else:
            render_bolin_optimized(ws, start_dt, end_dt, rows, final_budget_val, prod_cost)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
