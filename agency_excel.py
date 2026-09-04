# -*- coding: utf-8 -*-
"""
代理商 CUE Excel 渲染模組 (Agency Cue Excel Renderer)

以 openpyxl 精緻復刻三種代理商版型（2008傳媒 / 佳聖 / 凱絡），每平台一個工作表。
字級、列高、框線層次、底色、格式代碼、圖片錨點、頁首頁尾字串皆逐格對齊原始範例檔實測值。

- 字型一律「微軟正黑體」；版面靠 fitToWidth 縮放，不靠縮小字級塞版。
- 金額：2008 與凱絡三層價用會計 $ 格式；佳聖 金額與凱絡費用區用無 $ 純數字。
- 為求 LibreOffice/Excel 轉 PDF 穩定，數字一律寫入計算後實值（非公式）。
"""
import os
import calendar
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

import config
import agency_cue as ac

FONT = config.FONT_MAIN

# 數字格式
ACCT = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'   # 會計 $（2008、凱絡三層價）
NUM = "#,##0_);[Red](#,##0)"          # 無 $ 純數字（佳聖 金額、凱絡費用）
DAY_FMT = "0_ "                        # 每日檔次（尾隨空格）
DAY_RED = "#,##0_);[Red](#,##0)"      # 2008 每日合計
SUM_FMT = "#,##0_);(#,##0)"           # 2008 合計 F
MATERIAL_FMT = 'm"月"d"日"'            # 佳聖 素材日期
CARAT_H_FMT = "0_);[Red]\\(0\\)"      # 凱絡 檔數
CN_WD = "一二三四五六日"
EN_WD = "MTWTFSS"

YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")   # 週末亮黃


# =============================================================================
# 共用工具
# =============================================================================
def _cn_weekday(d):
    return CN_WD[d.weekday()]


def _en_weekday(d):
    return EN_WD[d.weekday()]


def _set(ws, coord, value, size=12, bold=False, align="center", valign="center",
         wrap=False, fmt=None, fill=None):
    c = ws[coord]
    c.value = value
    c.font = Font(name=FONT, size=size, bold=bold)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


def _merge(ws, rng):
    ws.merge_cells(rng)


def _edge(ws, coord, top=None, bottom=None, left=None, right=None):
    """疊加單元格邊框（不覆蓋既有其他邊）。"""
    c = ws[coord]
    b = c.border
    c.border = Border(
        left=Side(style=left) if left else b.left,
        right=Side(style=right) if right else b.right,
        top=Side(style=top) if top else b.top,
        bottom=Side(style=bottom) if bottom else b.bottom,
    )


def _box(ws, r1, c1, r2, c2, edge="medium", inner="thin"):
    """畫外框 edge、內線 inner 的方框（inner=None 則只有外框）。"""
    es = Side(style=edge)
    ins = Side(style=inner) if inner else Side(style=None)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(
                left=es if c == c1 else ins,
                right=es if c == c2 else ins,
                top=es if r == r1 else ins,
                bottom=es if r == r2 else ins,
            )


def _net_cell(ws, coord, net_display, size, bold=False, fmt=ACCT, align="center", fill=None):
    """實收欄：數字用指定金額格式，字串（專案回饋/計價於量販）用文字。"""
    if isinstance(net_display, str):
        _set(ws, coord, net_display, size=size, bold=bold, align=align, fill=fill)
    else:
        _set(ws, coord, net_display, size=size, bold=bold, fmt=fmt, align=align, fill=fill)


def _page(ws, orientation="landscape", margins=(0.31, 0.31, 1.1, 0.75)):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    l, r, t, b = margins
    ws.page_margins = PageMargins(left=l, right=r, top=t, bottom=b)


def _write_day_header(ws, start_col, start_dt, days, r_month, r_date, r_wd,
                      weekday_fn, month_style="en", size=12, date_as_datetime=False,
                      date_fmt="d", weekend_fill_rows=()):
    """
    寫日期表頭三列：月份(同月合併)、日號、星期。
    weekend_fill_rows: 需在週六日填黃底的列號集合。
    """
    from datetime import timedelta
    # 月份列（同月合併）
    seg_start = 0
    for i in range(1, days + 1):
        cur = start_dt + timedelta(days=i) if i < days else None
        prev = start_dt + timedelta(days=i - 1)
        boundary = (i == days) or (cur.month != prev.month)
        if boundary:
            c0 = start_col + seg_start
            c1 = start_col + i - 1
            mon = (start_dt + timedelta(days=seg_start)).month
            label = calendar.month_abbr[mon] if month_style == "en" else f"{mon}月"
            _set(ws, f"{get_column_letter(c0)}{r_month}", label, size=size, bold=(month_style == "en"),
                 align="left" if month_style == "en" else "center")
            if c1 > c0:
                _merge(ws, f"{get_column_letter(c0)}{r_month}:{get_column_letter(c1)}{r_month}")
            seg_start = i
    # 日號 + 星期
    for i in range(days):
        d = start_dt + timedelta(days=i)
        col = get_column_letter(start_col + i)
        wend = d.weekday() >= 5
        f_date = YELLOW if (wend and r_date in weekend_fill_rows) else None
        f_wd = YELLOW if (wend and r_wd in weekend_fill_rows) else None
        if date_as_datetime:
            _set(ws, f"{col}{r_date}", d, size=size, bold=True, fmt=date_fmt, fill=f_date)
        else:
            _set(ws, f"{col}{r_date}", d.day, size=size, bold=True, fmt=date_fmt, fill=f_date)
        _set(ws, f"{col}{r_wd}", weekday_fn(d), size=size, fill=f_wd)


def _period_dot(a, b):
    return f"{a.strftime('%Y.%m.%d')}-{b.strftime('%Y.%m.%d')}"


def _period_short(a, b):
    return f"{a.month}/{a.day}-{b.month}/{b.day}"


def _mat_str(d):
    return f"{d.month}/{d.day}" if d else ""


# =============================================================================
# 代理商 Logo（真品，隨專案打包）
# =============================================================================
_ASSET_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets")
LOGO_2008 = os.path.join(_ASSET_DIR, "logo_2008.png")     # 150×77
LOGO_DDRIVE = os.path.join(_ASSET_DIR, "logo_ddrive.png")  # 145×176


def _add_logo(ws, path, anchor_cell, width, height):
    if not os.path.exists(path):
        return
    img = XLImage(path)
    img.width = width
    img.height = height
    img.anchor = anchor_cell
    ws.add_image(img)


def _group_main_comp(rows):
    """main(+緊接 comp) 為一組；其餘各自一組。"""
    groups = []
    i = 0
    while i < len(rows):
        r = rows[i]
        if r["kind"] == ac.KIND_MAIN and i + 1 < len(rows) and rows[i + 1]["kind"] == ac.KIND_COMP:
            groups.append([r, rows[i + 1]])
            i += 2
        else:
            groups.append([r])
            i += 1
    return groups


# =============================================================================
# 2008傳媒版型
# =============================================================================
def _render_2008(wb, sheet, model, made_date):
    days = (model["end_date"] - model["start_date"]).days + 1
    start_dt = model["start_date"]
    sec = sheet["seconds"]
    ws = wb.create_sheet(title=_sheet_name(model, sheet))
    _page(ws, margins=(0.31, 0.31, 1.1, 0.75))

    # 第 1 欄(A)留白當左邊距；表格自 B 欄(X0=2)起，左外框補在 B 欄，與最右外框對稱。
    X0 = 2
    ws.column_dimensions[get_column_letter(1)].width = 3.2  # 左側留白欄（無框線）

    def cl(i):  # 邏輯欄 0-based → 欄字母（已含左留白位移）
        return get_column_letter(X0 + i)

    A_, B_, C_, D_, E_, F_, G_, H_ = (cl(i) for i in range(8))
    DAY0 = X0 + 8
    widths = [48.6, 22.0, 31.6, 33.6, 20.9, 32.6, 36.6, 29.4]
    for i, w in enumerate(widths):
        ws.column_dimensions[cl(i)].width = w
    for i in range(days):
        ws.column_dimensions[get_column_letter(DAY0 + i)].width = 8.9
    last_col = DAY0 + days - 1
    # 最後一欄（含跨月的 Sep）稍加寬，避免右側外框線被裁切遮住
    ws.column_dimensions[get_column_letter(last_col)].width = 11.5

    for r in range(1, 8):
        ws.row_dimensions[r].height = 40
    for r in (8, 9, 10):
        ws.row_dimensions[r].height = 41.9

    # Logo 右上（8.93cm × 4.85cm，96DPI 換算：寬 338px、高 183px）
    _add_logo(ws, LOGO_2008, f"{get_column_letter(max(DAY0, last_col - 5))}4", width=338, height=183)

    # 表頭 1~6：標籤靠左、值靠左
    header_pairs = [
        ("Client", model["client_name"]),
        ("Media Agency", "2008傳媒行銷股份有限公司"),
        ("Advertising Agency", ""),
        ("Product", model["product_name"]),
        ("Campaign", model.get("campaign", "")),
        ("Period", _period_dot(model["start_date"], model["end_date"])),
    ]
    for idx, (lab, val) in enumerate(header_pairs):
        rr = idx + 1
        _set(ws, f"{A_}{rr}", lab, size=28, bold=True, align="left")
        _set(ws, f"{C_}{rr}", val, size=28, bold=True, align="left")
        _merge(ws, f"{C_}{rr}:{H_}{rr}")

    # 欄位表頭 8~10（24pt 不粗）
    heads = [(A_, "媒體型態"), (B_, "地區"), (C_, "播出時段"), (D_, "定價"),
             (E_, "單位"), (F_, "次數"), (G_, "合計"), (H_, "素材\n提供時間")]
    for c, txt in heads:
        _set(ws, f"{c}8", txt, size=24, wrap=True)
        _merge(ws, f"{c}8:{c}10")
    _write_day_header(ws, DAY0, start_dt, days, 8, 9, 10, _cn_weekday, month_style="en",
                      size=24, date_as_datetime=True, date_fmt="d", weekend_fill_rows=(9, 10))

    day_cols = [(DAY0 + i, i) for i in range(days)]

    # 資料列
    r = 11
    groups = _group_main_comp(sheet["rows"])
    data_top = r
    schedule_rows = []
    data_rows = []  # [(row_number, row_dict)]：供左側欄延伸合併判斷
    total_spots = total_list = total_net = 0
    for grp in groups:
        gtop, gbot = r, r + len(grp) - 1
        main = grp[0]
        for k, row in enumerate(grp):
            rr = r + k
            data_rows.append((rr, row))
            ws.row_dimensions[rr].height = 105.65 if row["kind"] == ac.KIND_MAIN else 93.75
            _set(ws, f"{F_}{rr}", row["spots"], size=24, fmt=DAY_FMT)
            total_spots += row["spots"]
            if row["schedule"] is None:
                c0, c1 = get_column_letter(DAY0), get_column_letter(last_col)
                _set(ws, f"{c0}{rr}", row["spots"], size=24)
                if last_col > DAY0:
                    _merge(ws, f"{c0}{rr}:{c1}{rr}")
            else:
                for cidx, off in day_cols:
                    _set(ws, f"{get_column_letter(cidx)}{rr}", row["schedule"][off], size=24, fmt=DAY_FMT)
                schedule_rows.append((rr, row["schedule"]))
            if isinstance(row["net_display"], (int, float)):
                total_net += row["net_display"]
        _set(ws, f"{A_}{gtop}", main["media_label"], size=24, wrap=True)
        _set(ws, f"{B_}{gtop}", main["region_label"], size=24)
        _set(ws, f"{C_}{gtop}", main["daypart"], size=24)
        # 樂家康(超市)定價比照合計欄，顯示「計價於量販」而非 $0
        if main["kind"] in (ac.KIND_SUPER, ac.KIND_SUPER_REBATE):
            _set(ws, f"{D_}{gtop}", ac.NET_ON_MAG, size=24)
        else:
            _set(ws, f"{D_}{gtop}", main["list_total"], size=24, fmt=ACCT)
        _set(ws, f"{E_}{gtop}", f"{sec}秒", size=24)
        _net_cell(ws, f"{G_}{gtop}", main["net_display"], size=24)
        _set(ws, f"{H_}{gtop}", _mat_str(main["material"]), size=24, wrap=True)
        total_list += main["list_total"]
        # 定價/單位/合計：僅在同組（主+補償）內合併
        if gbot > gtop:
            for c in (D_, E_, G_):
                _merge(ws, f"{c}{gtop}:{c}{gbot}")
        r = gbot + 1

    data_bot = r - 1

    # 媒體型態/地區/播出時段：延伸合併蓋住「延續列」(media_label 為空的補償/回饋列)，
    # 使左側區塊整齊為單一高格。萬家福表每列各有媒體名，不會被併。
    blk_top = None
    for rr, row in data_rows:
        if row["media_label"]:
            if blk_top is not None and rr - 1 > blk_top:
                for c in (A_, B_, C_):
                    _merge(ws, f"{c}{blk_top}:{c}{rr - 1}")
            blk_top = rr
    if blk_top is not None and data_bot > blk_top:
        for c in (A_, B_, C_):
            _merge(ws, f"{c}{blk_top}:{c}{data_bot}")
    # 素材提供時間（H）：整張表同一素材日，整塊合併蓋住所有資料列（含量販→樂家康）
    if data_bot > data_top:
        _merge(ws, f"{H_}{data_top}:{H_}{data_bot}")
    # 合計列
    ws.row_dimensions[r].height = 80.15
    _set(ws, f"{B_}{r}", "合計", size=24)
    _merge(ws, f"{B_}{r}:{C_}{r}")
    _set(ws, f"{D_}{r}", total_list, size=24, fmt=ACCT)
    _set(ws, f"{F_}{r}", total_spots, size=24, fmt=SUM_FMT)
    _set(ws, f"{G_}{r}", total_net, size=24, fmt=ACCT)
    for cidx, off in day_cols:
        _set(ws, f"{get_column_letter(cidx)}{r}", sum(s[off] for _, s in schedule_rows),
             size=24, bold=True, fmt=DAY_RED)
    total_row = r

    # 框線層次：整表外框 double、日欄縱線 hair、表頭下緣 double、合計上緣 double
    _box(ws, 8, X0, total_row, last_col, edge="double", inner="hair")
    for c in range(X0, last_col + 1):
        _edge(ws, f"{get_column_letter(c)}10", bottom="double")   # 表頭區結束線
        _edge(ws, f"{get_column_letter(c)}{total_row}", top="double")  # 合計上緣

    # 費用區（四條細橫線的小表，無縱線）
    fr = total_row + 2
    f = sheet["fees"]
    fee_lines = [
        ("Budget (net)：", f["budget_net"]),
        (f"AC {int(f['ac_pct'])}%：", f["ac"]),
        ("5% Tax ：", f["tax"]),
        ("TOTAL ：", f["total"]),
    ]
    for i, (lab, val) in enumerate(fee_lines):
        rr = fr + i
        ws.row_dimensions[rr].height = 60
        _set(ws, f"{F_}{rr}", lab, size=24, bold=True, align="right")
        _net_cell(ws, f"{G_}{rr}", val, size=24, bold=True)
        _edge(ws, f"{F_}{rr}", top="thin", bottom="thin")
        _edge(ws, f"{G_}{rr}", top="thin", bottom="thin")

    # 備註 28pt 粗體
    rr = fr + len(fee_lines) + 1
    for line in model.get("remarks", []):
        ws.row_dimensions[rr].height = 55.75
        _set(ws, f"{A_}{rr}", line, size=28, bold=True, align="left")
        _merge(ws, f"{A_}{rr}:{get_column_letter(last_col)}{rr}")
        rr += 1
    return ws


# =============================================================================
# 佳聖 版型
# =============================================================================
def _render_ddrive(wb, sheet, model, made_date):
    days = (model["end_date"] - model["start_date"]).days + 1
    start_dt = model["start_date"]
    sec = sheet["seconds"]
    ws = wb.create_sheet(title=_sheet_name(model, sheet))
    # 左邊留白，列印 PDF 時左側不會切齊紙邊（YM 回饋）
    _page(ws, margins=(0.3, 0.2, 0.4, 0.2))

    # 列印頁首/頁尾
    ws.oddHeader.center.text = "佳聖媒體  戶外媒體排期表"
    ws.oddFooter.left.text = ("          群主管:_______________部主管:_______________"
                              "課主管:_______________承辦PM:_______________")
    ws.oddFooter.right.text = "佳聖媒體: ____________________          "

    # 第 1 欄(A)留白當左邊距；表格自 B 欄(X0=2)起，左外框補在 B 欄，與最右外框對稱。
    X0 = 2
    ws.column_dimensions[get_column_letter(1)].width = 3.2  # 左側留白欄（無框線）

    def cl(i):  # 邏輯欄 0-based → 欄字母（已含左留白位移）
        return get_column_letter(X0 + i)

    A_, B_, C_, D_, E_, F_, G_, H_ = (cl(i) for i in range(8))
    DAY0 = X0 + 8
    widths = [31.6, 25.5, 18.1, 19.9, 15.9, 21.4, 23.5, 23.5]
    for i, w in enumerate(widths):
        ws.column_dimensions[cl(i)].width = w
    for i in range(days):
        ws.column_dimensions[get_column_letter(DAY0 + i)].width = 8.5
    last_col = DAY0 + days - 1

    for r in (1, 2, 3, 4):
        ws.row_dimensions[r].height = 21.6
    for r in (5, 6, 7):
        ws.row_dimensions[r].height = 57
    for r in (9, 10, 11):
        ws.row_dimensions[r].height = 38

    _add_logo(ws, LOGO_DDRIVE, f"{A_}1", width=72, height=87)

    # 客戶/產品/刊期（標籤與值分開兩格；值前一空格；刊期短格式）
    _set(ws, f"{A_}5", "客戶：", size=22, bold=True, align="left")
    _set(ws, f"{B_}5", f" {model['client_name']}", size=22, bold=True, align="left")
    _set(ws, f"{A_}6", "產品：", size=22, bold=True, align="left")
    _set(ws, f"{B_}6", f" {model['product_name']}", size=22, bold=True, align="left")
    _set(ws, f"{A_}7", "刊期：", size=22, bold=True, align="left")
    _set(ws, f"{B_}7", f" {_period_short(model['start_date'], model['end_date'])}", size=22, bold=True, align="left")

    # 欄位表頭 9~11
    heads = [(A_, "媒體"), (B_, "地區"), (C_, "託播秒數"), (D_, "播出時段"),
             (E_, "次數"), (F_, "素材\n提供時間"), (G_, "定價\n(Net Cost)"), (H_, "專案執行價\n(Net Cost)")]
    for c, txt in heads:
        _set(ws, f"{c}9", txt, size=18, bold=True, wrap=True)
        _merge(ws, f"{c}9:{c}11")
    _write_day_header(ws, DAY0, start_dt, days, 9, 10, 11, _cn_weekday, month_style="cn", size=18)

    day_cols = [(DAY0 + i, i) for i in range(days)]
    r = 12
    data_top = r
    schedule_rows = []
    total_spots = total_list = total_net = 0
    is_wjf = sheet["platform"] == ac.PLATFORM_WJF
    first_data = r
    for row in sheet["rows"]:
        ws.row_dimensions[r].height = 62
        _set(ws, f"{A_}{r}", row["media_label"], size=18, wrap=True)
        _set(ws, f"{B_}{r}", row["region_label"], size=18)
        _set(ws, f"{C_}{r}", f"{sec}秒", size=18, wrap=True)
        _set(ws, f"{D_}{r}", row["daypart"], size=18, wrap=True)
        _set(ws, f"{E_}{r}", row["spots"], size=18)
        if row["material"]:
            _set(ws, f"{F_}{r}", row["material"], size=18, fmt=MATERIAL_FMT)
        is_super = row["kind"] in (ac.KIND_SUPER, ac.KIND_SUPER_REBATE)
        gval = ac.NET_ON_MAG if is_super else row["list_total"]
        _net_cell(ws, f"{G_}{r}", gval, size=18, fmt=NUM)
        _net_cell(ws, f"{H_}{r}", row["net_display"], size=18, fmt=NUM, align="right")
        total_spots += row["spots"]
        if not is_super and isinstance(row["list_total"], (int, float)):
            total_list += row["list_total"]
        if isinstance(row["net_display"], (int, float)):
            total_net += row["net_display"]
        if row["schedule"] is None:
            c0, c1 = get_column_letter(DAY0), get_column_letter(last_col)
            _set(ws, f"{c0}{r}", row["spots"], size=18)
            if last_col > DAY0:
                _merge(ws, f"{c0}{r}:{c1}{r}")
        else:
            for cidx, off in day_cols:
                _set(ws, f"{get_column_letter(cidx)}{r}", row["schedule"][off], size=16, fmt=DAY_FMT)
            schedule_rows.append((r, row["schedule"]))
        r += 1
    data_bot = r - 1

    # 平台專屬合併
    if is_wjf:
        # 萬家福：託播秒數/素材/專案執行價 合併量販+超市（前兩列）
        if data_bot >= first_data + 1:
            for c in (C_, F_, H_):
                _merge(ws, f"{c}{first_data}:{c}{first_data + 1}")
            _net_cell(ws, f"{H_}{first_data}", sheet["budget"] if not sheet["is_rebate_wave"] else ac.NET_REBATE,
                      size=18, fmt=NUM, align="right")
    else:
        # 全家：媒體/地區/秒數/時段/素材 直向合併整塊
        if data_bot > first_data:
            for c in (A_, B_, C_, D_, F_):
                _merge(ws, f"{c}{first_data}:{c}{data_bot}")

    # 小計列
    ws.row_dimensions[r].height = 41
    _set(ws, f"{A_}{r}", "小計", size=18)
    _merge(ws, f"{A_}{r}:{D_}{r}")
    _set(ws, f"{E_}{r}", total_spots, size=18, fmt="#,##0")
    _net_cell(ws, f"{G_}{r}", total_list, size=18, fmt=NUM)
    # H = 實收（僅主檔）
    net_main = 0
    for row in sheet["rows"]:
        if row["kind"] == ac.KIND_MAIN and isinstance(row["net_display"], (int, float)):
            net_main += row["net_display"]
    _net_cell(ws, f"{H_}{r}", net_main, size=18, fmt=NUM, align="right")
    for cidx, off in day_cols:
        _set(ws, f"{get_column_letter(cidx)}{r}", sum(s[off] for _, s in schedule_rows), size=16, fmt="#,##0")
    subtotal_row = r

    # 框線：外框 medium、內線 thin；小計底 medium
    _box(ws, 9, X0, subtotal_row, last_col, edge="medium", inner="thin")

    # 備註（左）＋ 請款；多月份請款每月一列往下排，避免長文字擠到右側費用欄
    br = subtotal_row + 2
    ws.row_dimensions[br].height = 52.7
    _set(ws, f"{A_}{br}", "備註：", size=22, align="left")
    pn = model.get("payment_note", "")
    pay_parts = pn.split("、") if pn else [""]
    for idx, part in enumerate(pay_parts):
        rr = br + 1 + idx
        if idx > 0:
            ws.row_dimensions[rr].height = 40
            # 續月縮排＝前綴「* 請款金額：」寬度(≈6全形)，對齊首列「8月份」
            part = "　　　　　　" + part
        _set(ws, f"{A_}{rr}", part, size=22, align="left", wrap=True)
        _merge(ws, f"{A_}{rr}:{E_}{rr}")

    # 費用框（標籤 / 值，中間留空，無框線，無 $）
    f = sheet["fees"]
    fee_lines = [("Total Net Cost", f["net"]), ("VAT   (5%)", f["vat"]), ("Total Gross Cost", f["gross"])]
    for i, (lab, val) in enumerate(fee_lines):
        rr = subtotal_row + 2 + i
        _set(ws, f"{F_}{rr}", lab, size=22, bold=True, align="left")
        _net_cell(ws, f"{H_}{rr}", val, size=22, bold=True, fmt=NUM, align="right")
    return ws


# =============================================================================
# 凱絡版型
# =============================================================================
def _render_carat(wb, sheet, model, made_date):
    from datetime import timedelta
    days = (model["end_date"] - model["start_date"]).days + 1
    start_dt = model["start_date"]
    sec = sheet["seconds"]
    ws = wb.create_sheet(title=_sheet_name(model, sheet))
    # 左邊留白，列印 PDF 時左側不會切齊紙邊（YM 回饋）
    _page(ws, margins=(0.3, 0.2, 0.3, 0.2))

    # 第 1 欄(A)留白當左邊距；表格自 B 欄(X0=2)起，左外框補在 B 欄，與最右外框對稱。
    X0 = 2
    ws.column_dimensions[get_column_letter(1)].width = 3.2  # 左側留白欄（無框線）

    def cl(i):  # 邏輯欄 0-based → 欄字母（已含左留白位移）
        return get_column_letter(X0 + i)

    MEDIA, REGION, DAYPART, MATERIAL, LIST_, MARKET, UNI, SPOTS, TOTAL, PROJ = (cl(i) for i in range(10))
    DAY0 = X0 + 10
    # I(總價)加寬到 17：左邊留白後 fitToWidth 會縮版，14.5 會讓 7 位數總價顯示 ######
    widths = [21.5, 18.7, 15.7, 9.5, 11.3, 11.5, 11.5, 10.5, 17.0, 14.1]
    for i, w in enumerate(widths):
        ws.column_dimensions[cl(i)].width = w
    for i in range(days):
        ws.column_dimensions[get_column_letter(DAY0 + i)].width = 8.6
    last_col = DAY0 + days - 1

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 30
    for r in (4, 5, 6, 7):
        ws.row_dimensions[r].height = 29.25

    made = made_date or model["start_date"]
    # 大標靠左、不合併
    _set(ws, f"{MEDIA}1", "凱絡媒體服務(股)公司廣播媒體排期表", size=16, bold=True, align="left")
    # 右上抬頭
    _set(ws, f"{TOTAL}2", "客   戶：", size=14, align="right")
    _set(ws, f"{PROJ}2", model["client_name"], size=14, align="left")
    _set(ws, f"{TOTAL}3", "產   品：", size=14, align="right")
    _set(ws, f"{PROJ}3", model["product_name"], size=14, align="left")
    _set(ws, f"{TOTAL}4", "日   期：", size=14, align="right")
    _set(ws, f"{PROJ}4", made.strftime("%Y/%m/%d"), size=14, align="left")
    _set(ws, f"{MEDIA}4", f"{start_dt.year}年{start_dt.month}月", size=14, align="left")

    # 欄位表頭 5~7
    heads = [(MEDIA, "媒體別"), (REGION, "地區"), (DAYPART, "時段"), (MATERIAL, "素材"),
             (LIST_, "定價\n(檔/Net)"), (MARKET, "市場價\n(檔/Net)"), (UNI, "統一價\n(檔/Net)"),
             (SPOTS, "檔數"), (TOTAL, "總價"), (PROJ, "專案價\n(Net)")]
    for c, txt in heads:
        _set(ws, f"{c}5", txt, size=12, bold=True, wrap=True)
        _merge(ws, f"{c}5:{c}7")
    _write_day_header(ws, DAY0, start_dt, days, 5, 6, 7, _en_weekday, month_style="en",
                      size=12, date_fmt="#,##0", weekend_fill_rows=(7,))
    # 最右加「總檔數」欄（對齊範本）
    tot_col = last_col + 1
    tl = get_column_letter(tot_col)
    ws.column_dimensions[tl].width = 9.5
    _set(ws, f"{tl}5", "總檔數", size=12, bold=True, wrap=True)
    _merge(ws, f"{tl}5:{tl}7")
    _box(ws, 5, X0, 7, tot_col, edge="medium", inner="thin")

    day_cols = [(DAY0 + i, i) for i in range(days)]
    r = 8
    data_top = r
    schedule_rows = []
    media_value = 0
    actual_net = sheet["fees"].get("subtotal") if isinstance(sheet["fees"].get("subtotal"), (int, float)) else 0

    # A 欄以（量販+超市）、（回饋量販+回饋超市）為組合併
    rows = sheet["rows"]
    groups = []
    i = 0
    while i < len(rows):
        if rows[i]["kind"] == ac.KIND_MAIN and i + 1 < len(rows) and rows[i + 1]["kind"] == ac.KIND_SUPER:
            groups.append([rows[i], rows[i + 1]]); i += 2
        elif rows[i]["kind"] == ac.KIND_REBATE and i + 1 < len(rows) and rows[i + 1]["kind"] == ac.KIND_SUPER_REBATE:
            groups.append([rows[i], rows[i + 1]]); i += 2
        else:
            groups.append([rows[i]]); i += 1

    for grp in groups:
        gtop = r
        for k, row in enumerate(grp):
            rr = r + k
            ws.row_dimensions[rr].height = 58
            _set(ws, f"{REGION}{rr}", row["region_label"], size=12, wrap=True)
            _set(ws, f"{DAYPART}{rr}", row["daypart"], size=12)
            _set(ws, f"{MATERIAL}{rr}", f'{sec}"CM', size=12, fmt="@")
            _set(ws, f"{LIST_}{rr}", row["list_per"], size=12, fmt=ACCT)
            _set(ws, f"{MARKET}{rr}", row["market_per"], size=12, fmt=ACCT)
            _set(ws, f"{UNI}{rr}", row["uni_per"], size=12, fmt=ACCT)
            _set(ws, f"{SPOTS}{rr}", row["spots"], size=12, fmt=CARAT_H_FMT)
            _set(ws, f"{TOTAL}{rr}", row["uni_total"], size=12, fmt=ACCT)
            # 回饋列在表上顯示「聲活回饋」（內部 net_display 仍為專案回饋）
            jval = row["net_display"]
            if row["kind"] == ac.KIND_REBATE and jval == ac.NET_REBATE:
                jval = ac.CARAT_REBATE_LABEL
            _net_cell(ws, f"{PROJ}{rr}", jval, size=12,
                      bold=(row["kind"] == ac.KIND_MAIN), align="center")
            media_value += row["uni_total"] if isinstance(row["uni_total"], (int, float)) else 0
            if row["schedule"] is None:
                c0, c1 = get_column_letter(DAY0), get_column_letter(last_col)
                _set(ws, f"{c0}{rr}", row["spots"], size=11)
                if last_col > DAY0:
                    _merge(ws, f"{c0}{rr}:{c1}{rr}")
            else:
                for cidx, off in day_cols:
                    _set(ws, f"{get_column_letter(cidx)}{rr}", row["schedule"][off], size=11, fmt=DAY_FMT)
                schedule_rows.append((rr, row["schedule"]))
            # 總檔數（最右欄）
            _set(ws, f"{tl}{rr}", row["spots"], size=12, fmt=CARAT_H_FMT)
        _set(ws, f"{MEDIA}{gtop}", grp[0]["media_label"], size=12, wrap=True)
        r += len(grp)
    data_bot = r - 1

    # 媒體別整塊合併為單一格（全表同一媒體別）；
    # 全家表另把地區/時段/素材併入回饋等延續列，使左側整齊。
    # 萬家福表量販/超市地區時段不同，故逐列保留、只合併媒體別。
    is_wjf = sheet["platform"] == ac.PLATFORM_WJF
    if data_bot > data_top:
        _merge(ws, f"{MEDIA}{data_top}:{MEDIA}{data_bot}")
        if not is_wjf:
            for c in (REGION, DAYPART, MATERIAL):
                _merge(ws, f"{c}{data_top}:{c}{data_bot}")

    _box(ws, data_top, X0, data_bot, tot_col, edge="medium", inner="thin")

    # 媒體總價值 / 優惠總價值（標籤＋數字，medium 方框）
    mv_top = data_bot + 1
    _set(ws, f"{MEDIA}{mv_top}", "媒體總價值(NET)", size=12, bold=True, align="left")
    _set(ws, f"{REGION}{mv_top}", media_value, size=12, fmt=ACCT)
    _set(ws, f"{MEDIA}{mv_top + 1}", "優惠總價值(NET)", size=12, bold=True, align="left")
    _set(ws, f"{REGION}{mv_top + 1}", media_value - actual_net, size=12, fmt=ACCT)
    _box(ws, mv_top, X0, mv_top + 1, X0 + 1, edge="medium", inner="thin")

    # 費用區（純數字；只有 Grand-Total 上細線下雙線）
    f = sheet["fees"]
    ac_txt = "-" if f.get("ac_free") else f.get("ac", 0)
    fee_lines = [("Sub-Total", f["subtotal"]), ("A.C     3%", ac_txt),
                 ("VAT    5%", f["vat"]), ("Grand-Total", f["grand"])]
    for i, (lab, val) in enumerate(fee_lines):
        rr = data_bot + 1 + i
        _set(ws, f"{TOTAL}{rr}", lab, size=12, bold=True, align="left")
        _net_cell(ws, f"{PROJ}{rr}", val, size=12, bold=True, fmt=NUM, align="right")
        if lab == "Grand-Total":
            _edge(ws, f"{TOTAL}{rr}", top="thin", bottom="double")
            _edge(ws, f"{PROJ}{rr}", top="thin", bottom="double")

    fee_bottom = data_bot + len(fee_lines)

    # 簽核帶：獨立 medium 方框、上下留白，四個簽核欄位垂直置中（不貼上下邊線，並保留簽章空間）
    sig_top = fee_bottom + 2                # 與費用區間隔一列
    sig_bottom = sig_top + 2               # 三列高，給簽章與上下留白
    for rr in range(sig_top, sig_bottom + 1):
        ws.row_dimensions[rr].height = 22
    sig_mid = sig_top + 1
    _set(ws, f"{MEDIA}{sig_mid}", "部主管：_________ ", size=12, align="left", valign="center")
    _set(ws, f"{MATERIAL}{sig_mid}", "課主管：_________ ", size=12, align="left", valign="center")
    _set(ws, f"{TOTAL}{sig_mid}", "媒體窗口：_________", size=12, align="left", valign="center")
    _set(ws, f"{get_column_letter(DAY0 + 4)}{sig_mid}", "承辦PM：_________", size=12, align="left", valign="center")
    _box(ws, sig_top, X0, sig_bottom, tot_col, edge="medium", inner=None)

    # 備註：medium 外框；上下各留一列 padding 使文字不貼邊線，內容列之間再空一列。
    # 「備／註」二字置中；標籤格與內容以直線分隔（備註格自帶框線）。
    rmk = model.get("remarks", [])
    rk_top = sig_bottom + 1
    lines_n = max(1, len(rmk))
    inner_span = lines_n * 2 - 1           # 內容列與空白列交錯，末列不留空白
    row_span = 1 + inner_span + 1          # 頂／底 padding 各一列
    rk_bot = rk_top + row_span - 1
    ws.row_dimensions[rk_top].height = 8   # 頂部留白
    ws.row_dimensions[rk_bot].height = 8   # 底部留白
    a = ws[f"{MEDIA}{rk_top}"]
    a.value = "備\n註"
    a.font = Font(name=FONT, size=12, bold=True)
    # 置中：備／註 兩字靠在一起置於中間，不要被 distributed 拉到上下兩端
    a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if row_span > 1:
        _merge(ws, f"{MEDIA}{rk_top}:{MEDIA}{rk_bot}")
    for i, line in enumerate(rmk):
        rr = rk_top + 1 + i * 2            # +1 跳過頂部 padding 列
        ws.row_dimensions[rr].height = 20  # 內容列高，文字垂直置中不貼邊線
        _set(ws, f"{REGION}{rr}", line, size=12, align="left", valign="center", wrap=True)
        _merge(ws, f"{REGION}{rr}:{tl}{rr}")
        if i < len(rmk) - 1:
            ws.row_dimensions[rr + 1].height = 16  # 內容列之間的空白間距列
    _box(ws, rk_top, X0, rk_bot, tot_col, edge="medium", inner=None)
    # 備註標籤格（備／註）與內容以直線分隔（additive 疊加，不覆蓋其他格線）
    for rr in range(rk_top, rk_bot + 1):
        _edge(ws, f"{MEDIA}{rr}", right="medium")

    # 整張表左右外框連續包住（表頭 row5 → 備註底），與最右外框對稱（YM 回饋）。
    # 用 _edge 疊加（非 _box），才不會抹掉內部格線。
    left_letter = get_column_letter(X0)
    for rr in range(5, rk_bot + 1):
        _edge(ws, f"{left_letter}{rr}", left="medium")
        _edge(ws, f"{tl}{rr}", right="medium")
    return ws


# =============================================================================
# 對外主函式
# =============================================================================
def _sheet_name(model, sheet):
    a = model["start_date"].strftime("%m%d")
    b = model["end_date"].strftime("%m%d")
    if model["agency"] == "2008傳媒":
        wan = round(sheet["budget"] / 10000) if sheet["budget"] else 0
        plat = "全家" if sheet["platform"] == ac.PLATFORM_FAMILY else "萬家福"
        name = f"{a}-{b}-{plat}-{wan}萬-{sheet['seconds']}秒"
    else:
        name = f"{sheet['platform']} {a}-{b} {sheet.get('seconds','')}秒"
    for ch in ':\\/?*[]':
        name = name.replace(ch, "")
    return name[:31]


def generate_agency_excel(model, made_date=None):
    """依 model 產出代理商 Excel（每平台一個工作表），回傳 bytes。"""
    wb = Workbook()
    wb.remove(wb.active)
    agency = model["agency"]
    for sheet in model["sheets"]:
        if agency == "2008傳媒":
            _render_2008(wb, sheet, model, made_date)
        elif agency == "佳聖":
            _render_ddrive(wb, sheet, model, made_date)
        elif agency == "凱絡":
            _render_carat(wb, sheet, model, made_date)
        else:
            _render_2008(wb, sheet, model, made_date)
    if not wb.worksheets:
        wb.create_sheet(title="空")
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
