# -*- coding: utf-8 -*-
"""
簡易模式 (Simple / 一鍵 CUE)

目標：業務只需「選平台組合 + 輸入預算」，系統以聰明預設（全省、標準走期、
各秒數各出一版）自動產出對應 CUE，讓業務先看到範本、再回頭調走期與細節。

Phase 1：重用既有已驗證的計算與渲染管線
  - 子公司三組  → calculator.calculate_plan_data → excel_renderer.generate_excel_from_scratch
  - 代理商 2008/凱絡 → agency_cue.build_agency_model → agency_excel.generate_agency_excel
  對每個秒數各產一張 CUE（子公司 10/15/20/30；代理商依平台），提供逐張下載與 ZIP。

Phase 2（未做）：子公司改「一檔四分頁」+ 範本加強欄、家樂福→萬家福/樂家康改名、全面美化。
"""
import io
from datetime import date, timedelta

import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook

import config
from utils import get_remarks_text, safe_filename, html_escape
from calculator import calculate_plan_data
from simple_render import render_subsidiary_workbook

import agency_cue as ac
from agency_excel import generate_agency_excel
from data_loader import load_config_from_cloud, load_agency_pricing_from_cloud


# =============================================================================
# 平台組合定義
# =============================================================================
# 子公司組合：media = calculator 用的媒體 key（"全家廣播"/"新鮮視"/"家樂福"）
#   註：範本中「企頻」與「全家」皆指全家店內廣播＝內部 key "全家廣播"。
#   家樂福已改名萬家福/樂家康，Phase 2 會改內部 key，Phase 1 先只改顯示。
SUB_COMBOS = {
    "sub_qp_fv":  {"label": "① 企頻 ＋ 新鮮視",
                   "media": ["全家廣播", "新鮮視"], "seconds": [10, 15, 20, 30],
                   "medium": "全家企頻 / 新鮮視　專案"},
    "sub_qp_wjf": {"label": "② 全家 ＋ 萬家福．樂家康",
                   "media": ["全家廣播", "家樂福"], "seconds": [10, 15, 20, 30],
                   "medium": "全家企頻 / 萬家福‧樂家康　專案"},
    "sub_fv_wjf": {"label": "③ 新鮮視 ＋ 萬家福．樂家康",
                   "media": ["新鮮視", "家樂福"], "seconds": [10, 15, 20, 30],
                   "medium": "新鮮視 / 萬家福‧樂家康　專案"},
}

# 代理商組合：platform = "family"（全家單組）或 "wjf"（萬家福.樂家康單組）
AGENCY_COMBOS = {
    "ag_2008_fam":   {"label": "④ 2008傳媒 － 全家單組",
                      "agency": "2008傳媒", "platform": "family", "seconds": [10, 15, 20, 30]},
    "ag_2008_wjf":   {"label": "⑤ 2008傳媒 － 萬家福．樂家康單組",
                      "agency": "2008傳媒", "platform": "wjf", "seconds": [10, 15, 20]},
    "ag_carat_fam":  {"label": "⑥ 凱絡 － 全家單組",
                      "agency": "凱絡", "platform": "family", "seconds": [10, 15, 20, 30]},
    "ag_carat_wjf":  {"label": "⑦ 凱絡 － 萬家福．樂家康單組",
                      "agency": "凱絡", "platform": "wjf", "seconds": [10, 15, 20]},
}

BUDGET_PRESETS = [200000, 250000, 300000, 400000]   # 常用預算快捷（20/25/30/40萬）


# =============================================================================
# 子公司：組 config + 逐秒數產表
# =============================================================================
def _build_sub_config(media_keys, second, share_first):
    """組出 calculate_plan_data 需要的 config：全省、單一秒數 100%、兩平台預算佔比。"""
    shares = {media_keys[0]: share_first, media_keys[1]: 100 - share_first} if len(media_keys) == 2 \
        else {media_keys[0]: 100}
    cfg = {}
    for m in media_keys:
        cfg[m] = {
            "share": shares[m],
            "is_national": True,
            "regions": list(config.REGIONS_ORDER),
            "sec_shares": {second: 100},
        }
    return cfg


def _gen_subsidiary(combo, budget, start_dt, end_dt, share_first,
                    client, tax_id, product, sales, prod_cost,
                    pricing_db, sec_factors, store_counts_num):
    """回傳 {bytes, fname, summary}：一檔多分頁（各秒數一版，範本版面）。"""
    days = (end_dt - start_dt).days + 1
    remarks = get_remarks_text(start_dt - timedelta(days=5),
                               f"{start_dt.year - 1911}年{start_dt.month}月", end_dt)
    seconds_rows = {}
    summary = []
    for sec in combo["seconds"]:
        cfg = _build_sub_config(combo["media"], sec, share_first)
        rows, _tla, _logs = calculate_plan_data(
            cfg, float(budget), days, pricing_db, sec_factors,
            store_counts_num, list(config.REGIONS_ORDER),
        )
        seconds_rows[sec] = rows
        for m in combo["media"]:
            tot = sum(int(r["spots"]) for r in rows
                      if r["media"] == m and "超市" not in str(r.get("region", "")))
            summary.append({"秒數": f"{sec}秒", "平台": m, "每區檔次×區數合計": tot})
    xlsx_bytes = render_subsidiary_workbook(
        combo, seconds_rows, float(budget), float(prod_cost),
        start_dt, end_dt, client, tax_id, product, remarks,
    )
    fname = safe_filename(f"{combo['label']}_{int(budget)//10000}萬.xlsx")
    return {"bytes": xlsx_bytes, "fname": fname, "summary": summary}


# =============================================================================
# 代理商：組 fam/wjf cfg + 逐秒數產表
# =============================================================================
def _gen_agency(combo, budget, start_dt, end_dt, client, product, campaign, agency_pricing):
    """回傳 {bytes, fname, summary}：各秒數合併成一檔（每秒數一分頁）。"""
    material_due = start_dt - timedelta(days=7)
    made = date.today()
    ac_pct = config.AGENCY_AC_DEFAULT.get(combo["agency"]) or 0
    base_model = None
    all_sheets = []
    summary = []
    for sec in combo["seconds"]:
        if combo["platform"] == "family":
            fam_cfg = {"enabled": True, "seconds": sec, "share": 100,
                       "rebate_pct": 0, "spots_override": 0, "auto_rebate": True}
            wjf_cfg = {"enabled": False}
        else:
            fam_cfg = {"enabled": False}
            wjf_cfg = {"enabled": True, "seconds": sec, "share": 100,
                       "rebate_pct": 0, "mag_override": 0, "is_rebate_wave": False}
        model = ac.build_agency_model(
            combo["agency"], client, product, campaign,
            start_dt, end_dt, float(budget),
            fam_cfg, wjf_cfg, ac.COMP_MOVE50, material_due, ac_pct,
            agency_pricing=agency_pricing,
        )
        if base_model is None:
            base_model = model
        all_sheets.extend(model["sheets"])
        for sh in model["sheets"]:
            main = sum(int(r.get("spots", 0)) for r in sh.get("rows", []) if r.get("kind") == "main")
            summary.append({"秒數": f"{sec}秒", "分頁": sh.get("platform", ""), "主檔次": main})
    combined = dict(base_model)
    combined["sheets"] = all_sheets
    xlsx = generate_agency_excel(combined, made)
    xlsx_bytes = xlsx.getvalue() if hasattr(xlsx, "getvalue") else xlsx
    fname = safe_filename(f"{combo['label']}_{int(budget)//10000}萬.xlsx")
    return {"bytes": xlsx_bytes, "fname": fname, "summary": summary}


# =============================================================================
# Excel → 網頁表格預覽（讓業務直接在頁面上看到範本長相）
# =============================================================================
def _sheet_to_html(ws):
    merged = {}
    skip = set()
    for rng in ws.merged_cells.ranges:
        merged[(rng.min_row, rng.min_col)] = (rng.max_row - rng.min_row + 1,
                                              rng.max_col - rng.min_col + 1)
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                if (rr, cc) != (rng.min_row, rng.min_col):
                    skip.add((rr, cc))
    out = ['<table style="border-collapse:collapse;font-family:微軟正黑體;font-size:12px">']
    for r in range(1, ws.max_row + 1):
        out.append("<tr>")
        for c in range(1, ws.max_column + 1):
            if (r, c) in skip:
                continue
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                v = ""
            elif isinstance(v, float):
                v = f"{v:,.0f}" if v == int(v) else f"{v:,.2f}"
            elif isinstance(v, int):
                v = f"{v:,}" if abs(v) >= 1000 else str(v)
            span = merged.get((r, c))
            rs = f' rowspan="{span[0]}"' if span and span[0] > 1 else ""
            cs = f' colspan="{span[1]}"' if span and span[1] > 1 else ""
            style = "border:1px solid #d0d0d0;padding:2px 6px;text-align:center;white-space:nowrap"
            if cell.font and cell.font.bold:
                style += ";font-weight:700"
            try:
                if cell.fill and cell.fill.patternType == "solid":
                    rgb = cell.fill.fgColor.rgb
                    if isinstance(rgb, str) and len(rgb) == 8 and rgb[2:] not in ("000000", "FFFFFF"):
                        style += f";background:#{rgb[2:]}"
            except Exception:
                pass
            out.append(f'<td{rs}{cs} style="{style}">{html_escape(str(v))}</td>')
        out.append("</tr>")
    out.append("</table>")
    return "".join(out)


def _workbook_to_html(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    return [(ws.title, _sheet_to_html(ws)) for ws in wb.worksheets]


@st.cache_data(ttl=600, show_spinner=False)
def _generate(combo_key, budget, start_iso, end_iso, share_first,
              client, tax_id, product, sales, prod_cost, campaign):
    """依條件產生 CUE（含 Excel bytes、各分頁 HTML 預覽、檔次摘要）；同條件走快取不重算。"""
    start_dt = date.fromisoformat(start_iso)
    end_dt = date.fromisoformat(end_iso)
    if combo_key in SUB_COMBOS:
        _, scn, pdb, sf, _sm, _err = load_config_from_cloud(config.GSHEET_SHARE_URL)
        out = _gen_subsidiary(SUB_COMBOS[combo_key], budget, start_dt, end_dt, share_first,
                              client, tax_id, product, sales, prod_cost, pdb, sf, scn)
    else:
        ap = load_agency_pricing_from_cloud(config.GSHEET_SHARE_URL)
        out = _gen_agency(AGENCY_COMBOS[combo_key], budget, start_dt, end_dt,
                          client, product, campaign, ap)
    out["preview"] = _workbook_to_html(out["bytes"])
    return out


# =============================================================================
# Streamlit UI
# =============================================================================
def render_simple_cue(store_counts_num, pricing_db, sec_factors, regions_order, sales_map=None):
    st.title("⚡ 簡易模式（一鍵產 CUE）")
    st.caption("只要選平台組合＋輸入預算，系統以標準走期與全省投放自動產出各秒數版本 CUE；先看範本，再回頭調走期與細節。")

    # --- 平台組合 ---
    all_combos = {**{k: v["label"] for k, v in SUB_COMBOS.items()},
                  **{k: v["label"] for k, v in AGENCY_COMBOS.items()}}
    combo_key = st.radio("平台組合", list(all_combos.keys()),
                         format_func=lambda k: all_combos[k])
    is_sub = combo_key in SUB_COMBOS

    # --- 預算：自由輸入 + 常用快捷 ---
    st.markdown("**預算（未稅 Net）**")
    if "simple_budget" not in st.session_state:
        st.session_state["simple_budget"] = 250000
    pcols = st.columns(len(BUDGET_PRESETS) + 1)
    for i, amt in enumerate(BUDGET_PRESETS):
        if pcols[i].button(f"{amt // 10000}萬", key=f"preset_{amt}"):
            st.session_state["simple_budget"] = amt
    budget = pcols[-1].number_input("自訂金額", min_value=0,
                                    value=int(st.session_state["simple_budget"]), step=10000,
                                    label_visibility="collapsed")
    st.session_state["simple_budget"] = int(budget)

    # --- 走期：開始日 + 週數 ---
    c1, c2, c3 = st.columns(3)
    with c1:
        start_dt = st.date_input("開始日", value=date.today())
    with c2:
        weeks = st.number_input("走期（週）", min_value=1, max_value=12, value=2, step=1)
    end_dt = start_dt + timedelta(days=int(weeks) * 7 - 1)
    with c3:
        st.text_input("結束日（自動）", value=end_dt.strftime("%Y-%m-%d"), disabled=True)

    # --- 客戶/產品/業務（可留白，先產範本）---
    with st.expander("客戶／產品資訊（可留白，先產範本再補）", expanded=False):
        cc1, cc2, cc3 = st.columns(3)
        client = cc1.text_input("客戶名稱", "")
        product = cc2.text_input("產品名稱", "")
        sales_options = list(sales_map.keys()) if sales_map else []
        sales = cc3.selectbox("業務", ["—"] + sales_options) if sales_options else ""
        sales = "" if sales == "—" else sales
        tax_id = cc1.text_input("統一編號", "")
        prod_cost = cc2.number_input("製作費（未稅）", min_value=0, value=0, step=1000)
        campaign = cc3.text_input("Campaign（2008 用）", "")

    # 子公司專屬：預算佔比（兩平台）
    share_first = 50
    if is_sub:
        media_keys = SUB_COMBOS[combo_key]["media"]
        share_first = st.slider(
            f"預算佔比：{_disp_media(media_keys[0])} ％（其餘給 {_disp_media(media_keys[1])}）",
            0, 100, 50, step=5)

    combo = SUB_COMBOS[combo_key] if is_sub else AGENCY_COMBOS[combo_key]

    st.divider()
    if budget <= 0:
        st.info("輸入預算金額後，範本會自動顯示於下方。")
        return

    # 條件選完即自動產生 + 即時預覽（同條件走快取，不重算）
    try:
        with st.spinner("產生範本中…"):
            out = _generate(combo_key, int(budget),
                            start_dt.isoformat(), end_dt.isoformat(), int(share_first),
                            client, tax_id, product, sales, int(prod_cost), campaign)
    except Exception as e:
        st.error(f"產生失敗：{e}")
        st.exception(e)
        return

    st.success(f"{all_combos[combo_key]}｜預算 {int(budget)//10000}萬｜走期 {start_dt}～{end_dt}"
               f"｜{len(combo['seconds'])} 個秒數版本（客戶挑秒數）")
    st.download_button("⬇ 下載此 CUE（Excel）", data=out["bytes"], file_name=out["fname"],
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary")

    # --- 即時範本預覽（每個秒數一個分頁）---
    st.markdown("#### 📄 範本預覽")
    preview = out.get("preview") or []
    if preview:
        tabs = st.tabs([title for title, _ in preview])
        for tab, (_title, html) in zip(tabs, preview):
            with tab:
                components.html(
                    f'<div style="overflow:auto">{html}</div>',
                    height=520, scrolling=True)
    if out.get("summary"):
        with st.expander("檔次摘要（各秒數）"):
            st.dataframe(out["summary"], use_container_width=True, hide_index=True)


def _disp_media(m):
    return {"全家廣播": "全家企頻", "家樂福": "萬家福‧樂家康"}.get(m, m)
