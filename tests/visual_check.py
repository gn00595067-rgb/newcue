# -*- coding: utf-8 -*-
"""
代理商 CUE 版面驗收：以與原始範例完全相同的參數產生三份 Excel，
並對關鍵儲存格值、logo、頁首頁尾、週末底色等做結構性斷言。

執行：py tests/visual_check.py [輸出資料夾]
產出的 xlsx 可再用 Excel/LibreOffice 轉 PDF 與原始範例併排目視比對。
"""
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import agency_cue as ac
from agency_excel import generate_agency_excel


def _fam(**k):
    d = {"enabled": True, "seconds": 15, "share": 100, "rebate_pct": 0, "spots_override": 0}
    d.update(k)
    return d


def _wjf(**k):
    d = {"enabled": True, "seconds": 20, "share": 100, "rebate_pct": 0, "mag_override": 0, "is_rebate_wave": False}
    d.update(k)
    return d


def _load(bytes_):
    import io
    return openpyxl.load_workbook(io.BytesIO(bytes_))


def _cells(ws):
    return {c.coordinate: c.value for row in ws.iter_rows() for c in row if c.value is not None}


def _has_yellow(ws, coord):
    fill = ws[coord].fill
    return fill is not None and fill.fgColor is not None and str(fill.fgColor.rgb).endswith("FFFF00")


def check():
    out_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), "_visual_out")
    os.makedirs(out_dir, exist_ok=True)
    fails = []

    def ok(cond, msg):
        (print(f"  ✓ {msg}") if cond else (print(f"  ✗ {msg}"), fails.append(msg)))

    # ---- A. 2008 全家 ----
    print("[2008 全家] 2026/08/05~09/01 25萬 15秒 補償原50% AC3%")
    m = ac.build_agency_model("2008傳媒", "統一企業", "統一木瓜牛乳", "統一木瓜牛乳",
                              date(2026, 8, 5), date(2026, 9, 1), 250000,
                              _fam(seconds=15), None, ac.COMP_MOVE50, date(2026, 7, 24), 3.0)
    b = generate_agency_excel(m)
    open(os.path.join(out_dir, "2008_family.xlsx"), "wb").write(b)
    wb = _load(b); ws = wb.worksheets[0]; cv = _cells(ws)
    ok(cv.get("A1") == "Client" and ws["A1"].alignment.horizontal == "left", "表頭 Client 靠左")
    ok(cv.get("D11") == 2808000, "定價 2,808,000")
    ok(cv.get("F11") == 960, "主檔次 960")
    ok(cv.get("G18") == 270375, "TOTAL 270,375")
    ok(len(ws._images) == 1, "有 1 個 logo")
    # 週末黃底（8/9 為週六→日期列與星期列黃）
    ok(_has_yellow(ws, "M9") or _has_yellow(ws, "M10"), "週末日期/星期黃底")
    ok(any("※" in str(v) for v in cv.values()), "備註 ※ 原文")

    # ---- C. 佳聖 萬家福 ----
    print("[佳聖 萬家福] 2026/08/08~08/27 12.5萬 20秒")
    m = ac.build_agency_model("佳聖", "統一企業-飼料部", "元氣御選", "",
                              date(2026, 8, 8), date(2026, 8, 27), 125000,
                              None, _wjf(seconds=20), ac.COMP_NONE, date(2026, 8, 1), None)
    b = generate_agency_excel(m)
    open(os.path.join(out_dir, "ddrive_wjf.xlsx"), "wb").write(b)
    wb = _load(b); ws = wb.worksheets[0]; cv = _cells(ws)
    ok(cv.get("A12") == "萬家福" and cv.get("A13") == "樂家康", "媒體 萬家福/樂家康")
    ok(cv.get("B12") == "量販" and cv.get("B13") == "超市", "地區 量販/超市")
    ok(cv.get("D12") == "09-23" and cv.get("D13") == "00-24", "時段 09-23/00-24")
    ok(cv.get("G12") == 840000, "定價 840,000")
    ok(cv.get("G13") == ac.NET_ON_MAG, "超市 計價於量販")
    ok("佳聖媒體" in (ws.oddHeader.center.text or ""), "頁首 佳聖媒體")
    ok("承辦PM" in (ws.oddFooter.left.text or ""), "頁尾簽核")
    ok(len(ws._images) == 1, "有 M Drive logo")
    ok(ws["G12"].number_format.startswith("#,##0"), "定價無 $ 純數字格式")

    # ---- D. 凱絡 萬家福 ----
    print("[凱絡 萬家福] 2026/08/17~09/01 25萬 15秒 回饋10% A.C免收")
    m = ac.build_agency_model("凱絡", "客戶", "產品", "",
                              date(2026, 8, 17), date(2026, 9, 1), 250000,
                              None, _wjf(seconds=15, rebate_pct=10), ac.COMP_NONE, date(2026, 8, 10), None)
    b = generate_agency_excel(m)
    open(os.path.join(out_dir, "carat_wjf.xlsx"), "wb").write(b)
    wb = _load(b); ws = wb.worksheets[0]; cv = _cells(ws)
    ok(cv.get("A1") == "凱絡媒體服務(股)公司廣播媒體排期表" and ws["A1"].alignment.horizontal == "left", "大標靠左")
    ok(cv.get("I8") == 1428000, "量販總價 1,428,000")
    ok(cv.get("J8") == 250000, "專案價 250,000")
    ok(any("媒體總價值" in str(k) and cv.get("B" + k[1:]) == 4263600 for k in cv if str(cv[k]).startswith("媒體總價值")) or cv.get("B12") == 4263600 or cv.get("B13") == 4263600, "媒體總價值 4,263,600")
    ok(any(str(v).startswith("1.更改媒體排期表") for v in cv.values()), "凱絡備註原文")
    # 週末：只有星期列黃、日期列不黃（8/23 週六）
    # 找出星期列(第7列)週六欄
    wd_yellow = any(_has_yellow(ws, f"{openpyxl.utils.get_column_letter(c)}7") for c in range(11, ws.max_column + 1))
    date_yellow = any(_has_yellow(ws, f"{openpyxl.utils.get_column_letter(c)}6") for c in range(11, ws.max_column + 1))
    ok(wd_yellow and not date_yellow, "凱絡僅星期列週末黃底")

    print()
    if fails:
        print(f"❌ {len(fails)} 項未通過")
        return 1
    print(f"✅ 全部通過，Excel 已輸出至 {out_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(check())
